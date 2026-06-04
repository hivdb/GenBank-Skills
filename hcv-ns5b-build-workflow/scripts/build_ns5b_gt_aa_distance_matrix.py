#!/usr/bin/env python3

from __future__ import annotations

import argparse
import subprocess
from pathlib import Path

from openpyxl import Workbook


FASTA_EXTENSIONS = {".fa", ".fasta", ".faa"}
MISSING_AA = {"-", "X", "?", "*"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Multiple-align NS5B genotype amino-acid consensus FASTA, slice aligned "
            "positions 150-321, and write pairwise genotype amino-acid distances."
        )
    )
    parser.add_argument("--input-fasta", default="outputs/NS5B_GT_Consensus.fasta")
    parser.add_argument("--aligned-fasta", default="outputs/NS5B_GT_Consensus_aligned.fasta")
    parser.add_argument("--output-xlsx", default="outputs/NS5B_GT_AA_Distance_Pos150_321.xlsx")
    parser.add_argument(
        "--details-xlsx",
        default="temp/ns5b_gt_aa_distance_matrix/NS5B_GT_AA_Distance_Pos150_321_details.xlsx",
        help="Workbook for supporting sheets that are not kept in the main output",
    )
    parser.add_argument("--start", type=int, default=150, help="1-based aligned start position")
    parser.add_argument("--end", type=int, default=321, help="1-based aligned end position, inclusive")
    parser.add_argument("--mafft-bin", default="mafft")
    return parser.parse_args()


def read_fasta(path: Path) -> list[tuple[str, str]]:
    records: list[tuple[str, str]] = []
    name: str | None = None
    chunks: list[str] = []
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if name is not None:
                    records.append((name, "".join(chunks)))
                name = line[1:].strip().split()[0]
                chunks = []
            else:
                chunks.append(line)
    if name is not None:
        records.append((name, "".join(chunks)))
    if not records:
        raise RuntimeError(f"No FASTA records found in {path}")
    return records


def write_fasta(path: Path, records: list[tuple[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for name, sequence in records:
            handle.write(f">{name}\n")
            for start in range(0, len(sequence), 70):
                handle.write(sequence[start : start + 70] + "\n")


def run_mafft(input_fasta: Path, aligned_fasta: Path, mafft_bin: str) -> None:
    aligned_fasta.parent.mkdir(parents=True, exist_ok=True)
    result = subprocess.run(
        [mafft_bin, "--auto", str(input_fasta)],
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    aligned_fasta.write_text(result.stdout, encoding="utf-8")


def pairwise_distance(seq_a: str, seq_b: str) -> tuple[float | None, int, int]:
    differences = 0
    compared = 0
    for aa_a, aa_b in zip(seq_a, seq_b):
        if aa_a.upper() in MISSING_AA or aa_b.upper() in MISSING_AA:
            continue
        compared += 1
        if aa_a != aa_b:
            differences += 1
    denominator = len(seq_a)
    if denominator == 0:
        return None, differences, compared
    return differences / denominator, differences, compared


def add_matrix_sheet(
    wb: Workbook,
    title: str,
    names: list[str],
    values: dict[tuple[str, str], float | int | None],
    number_format: str | None = None,
    column_names: list[str] | None = None,
) -> None:
    column_names = column_names or names
    ws = wb.create_sheet(title)
    ws.append(["Genotype", *column_names])
    for row_name in names:
        row: list[str | float | int | None] = [row_name]
        for col_name in column_names:
            row.append(values[(row_name, col_name)])
        ws.append(row)
    ws.freeze_panes = "B2"
    if number_format:
        for row in ws.iter_rows(min_row=2, min_col=2):
            for cell in row:
                cell.number_format = number_format


def create_workbook() -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    return wb


def write_workbook(
    output_path: Path,
    details_path: Path,
    aligned_records: list[tuple[str, str]],
    window_records: list[tuple[str, str]],
    start: int,
    end: int,
    input_fasta: Path,
    aligned_fasta: Path,
) -> None:
    names = [name for name, _ in window_records]
    seqs = dict(window_records)
    distances: dict[tuple[str, str], float | None] = {}
    differences: dict[tuple[str, str], int] = {}
    compared_positions: dict[tuple[str, str], int] = {}

    for name_a in names:
        for name_b in names:
            distance, difference_count, compared = pairwise_distance(seqs[name_a], seqs[name_b])
            distances[(name_a, name_b)] = distance
            differences[(name_a, name_b)] = difference_count
            compared_positions[(name_a, name_b)] = compared

    name_order = {name: index for index, name in enumerate(names)}
    display_distances = {
        (name_a, name_b): (
            distances[(name_a, name_b)]
            if name_order[name_a] < name_order[name_b]
            else None
        )
        for name_a in names
        for name_b in names
    }

    wb = create_workbook()
    add_matrix_sheet(wb, "distance_matrix", names[:-1], display_distances, "0.0%", column_names=names[1:])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    details_wb = create_workbook()
    add_matrix_sheet(details_wb, "difference_count", names, differences)
    add_matrix_sheet(details_wb, "compared_positions", names, compared_positions)

    ws = details_wb.create_sheet("aligned_window")
    ws.append(["Genotype", "AlignedPositions", "AASequence"])
    for name, sequence in window_records:
        ws.append([name, f"{start}-{end}", sequence])
    ws.freeze_panes = "A2"

    meta = details_wb.create_sheet("metadata")
    meta_rows = [
        ("input_fasta", str(input_fasta.resolve())),
        ("aligned_fasta", str(aligned_fasta.resolve())),
        ("output_xlsx", str(output_path.resolve())),
        ("details_xlsx", str(details_path.resolve())),
        ("aligned_position_start_1based", start),
        ("aligned_position_end_1based_inclusive", end),
        ("window_length_columns", end - start + 1),
        ("distance_definition", "AA differences / full aligned window length"),
        ("ignored_characters", "".join(sorted(MISSING_AA))),
        ("record_count", len(aligned_records)),
    ]
    for row in meta_rows:
        meta.append(row)

    details_path.parent.mkdir(parents=True, exist_ok=True)
    details_wb.save(details_path)


def main() -> int:
    args = parse_args()
    input_fasta = Path(args.input_fasta)
    aligned_fasta = Path(args.aligned_fasta)
    output_xlsx = Path(args.output_xlsx)
    details_xlsx = Path(args.details_xlsx)
    if args.start < 1 or args.end < args.start:
        raise SystemExit("--start/--end must define a valid 1-based inclusive range")
    if input_fasta.suffix.lower() not in FASTA_EXTENSIONS:
        raise SystemExit(f"Input FASTA extension should be one of {sorted(FASTA_EXTENSIONS)}")

    run_mafft(input_fasta, aligned_fasta, args.mafft_bin)
    aligned_records = read_fasta(aligned_fasta)
    lengths = {len(sequence) for _, sequence in aligned_records}
    if len(lengths) != 1:
        raise RuntimeError(f"MAFFT output has inconsistent aligned lengths: {sorted(lengths)}")
    aligned_length = next(iter(lengths))
    if args.end > aligned_length:
        raise SystemExit(f"Requested end position {args.end} exceeds aligned length {aligned_length}")

    window_records = [
        (name, sequence[args.start - 1 : args.end])
        for name, sequence in aligned_records
    ]
    write_workbook(
        output_xlsx,
        details_xlsx,
        aligned_records,
        window_records,
        args.start,
        args.end,
        input_fasta,
        aligned_fasta,
    )
    print(f"aligned_fasta={aligned_fasta.resolve()}")
    print(f"output_xlsx={output_xlsx.resolve()}")
    print(f"details_xlsx={details_xlsx.resolve()}")
    print(f"record_count={len(aligned_records)}")
    print(f"aligned_length={aligned_length}")
    print(f"window_length={args.end - args.start + 1}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
