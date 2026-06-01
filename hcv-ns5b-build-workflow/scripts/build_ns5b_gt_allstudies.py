#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

# Historical output flow kept here for reference only.
# import csv


BLAST_OUTFMT = "6 qseqid sseqid length mismatch gaps pident evalue bitscore qstart qend sstart send"
REFERENCE_GTS = tuple(str(i) for i in range(1, 9))
RESISTANCE_POSITIONS = [150, 159, 206, 282, 316, 320, 321]
TARGET_GENE = "NS5B"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build per-study NS5B genotype distance reports by aligning FASTA sequences "
            "to GT1-GT8 NS5B nucleotide references."
        )
    )
    parser.add_argument("--excel-file", required=True, help="Path to the spreadsheet")
    parser.add_argument("--sheet", required=True, help="Worksheet name to read")
    parser.add_argument("--fasta-dir", required=True, help="Directory containing study FASTA files")
    parser.add_argument("--reference-fasta", required=True, help="Path to HCV_GT_RefSeqs.fasta")
    parser.add_argument("--output-dir", default="outputs", help="Base output directory")
    parser.add_argument("--refid-column", default="RefID")
    parser.add_argument("--refname-column", default="RefName")
    parser.add_argument("--numpatients-column", default="NumPts")
    parser.add_argument("--positive-column", action="append", default=["NS5BCount"])
    parser.add_argument("--min-aligned-nt", type=int, default=200, help="Skip hits shorter than this overlap length")
    return parser.parse_args()


def sanitize_label(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return text.strip("._-") or "job"


def script_temp_dir() -> Path:
    path = Path("temp") / Path(__file__).stem
    path.mkdir(parents=True, exist_ok=True)
    return path


def make_job_dir(base_output_dir: Path, excel_file: Path, sheet_name: str) -> Path:
    label = sanitize_label(f"{excel_file.stem}_{sheet_name}_ns5b_gt_distance")
    return Path(tempfile.mkdtemp(prefix=f"{label}_", dir=script_temp_dir()))


def parse_positive_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_fasta(path: Path) -> list[tuple[str, str]]:
    records: list[tuple[str, str]] = []
    header: str | None = None
    chunks: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line:
            continue
        if line.startswith(">"):
            if header is not None:
                records.append((header, "".join(chunks).upper()))
            header = line[1:].strip()
            chunks = []
        else:
            chunks.append(re.sub(r"\s+", "", line))
    if header is not None:
        records.append((header, "".join(chunks).upper()))
    return records


def accession_from_header(header: str) -> str:
    return header.split()[0]


def load_reference_ns5b(reference_fasta: Path) -> dict[str, tuple[str, str]]:
    refs: dict[str, tuple[str, str]] = {}
    for header, sequence in parse_fasta(reference_fasta):
        token = header.split()[0]
        match = re.match(r"HCV(\d+)(.+)", token)
        if not match:
            continue
        gt = match.group(1)
        gene = match.group(2)
        if gt in REFERENCE_GTS and gene == TARGET_GENE:
            refs[gt] = (header, sequence)
    missing = [gt for gt in REFERENCE_GTS if gt not in refs]
    if missing:
        raise RuntimeError(f"Missing {TARGET_GENE} references for GTs: {', '.join(missing)}")
    return refs


def load_filtered_studies(
    excel_file: Path,
    sheet_name: str,
    refid_column: str,
    refname_column: str,
    numpatients_column: str,
    positive_columns: list[str],
) -> list[dict[str, Any]]:
    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"Worksheet '{sheet_name}' was not found in {excel_file}")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError(f"Worksheet '{sheet_name}' is empty")
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    header_index = {header: idx for idx, header in enumerate(headers) if header}
    required = [refid_column, refname_column, numpatients_column, *positive_columns]
    for name in required:
        if name not in header_index:
            raise RuntimeError(f"Column '{name}' was not found in worksheet '{sheet_name}'")

    studies: list[dict[str, Any]] = []
    for row in rows[1:]:
        numpts = parse_positive_number(row[header_index[numpatients_column]])
        if numpts is None or numpts <= 0:
            continue
        valid = True
        positive_values: dict[str, float] = {}
        for name in positive_columns:
            value = parse_positive_number(row[header_index[name]])
            if value is None or value <= 0:
                valid = False
                break
            positive_values[name] = value
        if not valid:
            continue
        refid = str(row[header_index[refid_column]]).strip()
        refname = str(row[header_index[refname_column]]).strip()
        if not refid:
            continue
        studies.append(
            {
                "RefID": refid,
                "RefName": refname,
                numpatients_column: numpts,
                **positive_values,
            }
        )
    return studies


def filename_matches_refid(filename: str, refid: str) -> bool:
    return filename.startswith(f"{refid}_")


def find_study_fasta_files(fasta_dir: Path, studies: list[dict[str, Any]]) -> list[dict[str, Any]]:
    files = sorted(path for path in fasta_dir.rglob("*") if path.is_file())
    result: list[dict[str, Any]] = []
    for study in studies:
        matches = [path for path in files if filename_matches_refid(path.name, study["RefID"])]
        if matches:
            if len(matches) != 1:
                raise RuntimeError(f"Expected exactly one FASTA file for RefID {study['RefID']}, found {len(matches)}")
            result.append({**study, "fasta_path": str(matches[0])})
    return result


def write_fasta_entries(path: Path, entries: list[tuple[str, str]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for header, sequence in entries:
            handle.write(f">{header}\n")
            for start in range(0, len(sequence), 70):
                handle.write(sequence[start : start + 70] + "\n")


def build_reference_db(job_dir: Path, refs: dict[str, tuple[str, str]]) -> tuple[Path, dict[str, str]]:
    ref_fasta = job_dir / "ns5b_gt_refs.fasta"
    entries: list[tuple[str, str]] = []
    subject_id_to_gt: dict[str, str] = {}
    for gt in REFERENCE_GTS:
        subject_id = f"GT{gt}"
        subject_id_to_gt[subject_id] = gt
        entries.append((subject_id, refs[gt][1]))
    write_fasta_entries(ref_fasta, entries)
    db_prefix = job_dir / "ns5b_gt_refs_db"
    subprocess.run(
        [
            "makeblastdb",
            "-in",
            str(ref_fasta),
            "-dbtype",
            "nucl",
            "-out",
            str(db_prefix),
            "-parse_seqids",
        ],
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        text=True,
    )
    return db_prefix, subject_id_to_gt


def run_blastn_for_study(query_path: Path, db_prefix: Path) -> list[dict[str, Any]]:
    out_path = query_path.with_suffix(".blast.tsv")
    subprocess.run(
        [
            "blastn",
            "-query",
            str(query_path),
            "-db",
            str(db_prefix),
            "-dust",
            "no",
            "-task",
            "blastn",
            "-evalue",
            "1e-6",
            "-max_hsps",
            "1",
            "-max_target_seqs",
            "8",
            "-outfmt",
            BLAST_OUTFMT,
            "-out",
            str(out_path),
        ],
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        text=True,
    )
    hits: list[dict[str, Any]] = []
    for line in out_path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        p = line.split("\t")
        length = int(p[2])
        mismatch = int(p[3])
        gaps = int(p[4])
        hits.append(
            {
                "qseqid": p[0],
                "sseqid": p[1],
                "length": length,
                "mismatch": mismatch,
                "gaps": gaps,
                "pident": float(p[5]),
                "evalue": float(p[6]),
                "bitscore": float(p[7]),
                "qstart": int(p[8]),
                "qend": int(p[9]),
                "sstart": int(p[10]),
                "send": int(p[11]),
                "distance": (mismatch + gaps) / length if length else None,
            }
        )
    try:
        out_path.unlink()
    except OSError:
        pass
    return hits


def build_rows_for_study(
    study: dict[str, Any],
    db_prefix: Path,
    subject_id_to_gt: dict[str, str],
    min_aligned_nt: int,
) -> list[dict[str, Any]]:
    fasta_path = Path(study["fasta_path"])
    hits = run_blastn_for_study(fasta_path, db_prefix)
    best_by_query_gt: dict[tuple[str, str], dict[str, Any]] = {}
    for hit in hits:
        gt = subject_id_to_gt.get(hit["sseqid"])
        if gt is None:
            continue
        key = (hit["qseqid"], gt)
        current = best_by_query_gt.get(key)
        if current is None or (
            hit["distance"],
            -hit["length"],
            -hit["bitscore"],
        ) < (
            current["distance"],
            -current["length"],
            -current["bitscore"],
        ):
            best_by_query_gt[key] = hit

    rows: list[dict[str, Any]] = []
    for header, _sequence in parse_fasta(fasta_path):
        accession = accession_from_header(header)
        gt_hits: dict[str, dict[str, Any] | None] = {
            gt: best_by_query_gt.get((accession, gt)) for gt in REFERENCE_GTS
        }

        valid_hits = {
            gt: hit
            for gt, hit in gt_hits.items()
            if hit is not None and int(hit["length"]) >= min_aligned_nt
        }
        if not valid_hits:
            continue

        best_gt, best_hit = min(
            valid_hits.items(),
            key=lambda item: (
                item[1]["distance"],
                -item[1]["length"],
                -item[1]["bitscore"],
            ),
        )
        covered_positions = resistance_positions_covered(best_hit)

        row = {
            "RefID": study["RefID"],
            "RefName": study["RefName"],
            "GenBankAccession": accession,
            "BestGT": best_gt,
            "BestGTDistance": best_hit["distance"],
            "AlignedNT": best_hit["length"],
            "ContainsResistancePosition": "yes" if covered_positions else "no",
            "ResistancePositionsCovered": ",".join(str(pos) for pos in covered_positions),
        }
        for gt in REFERENCE_GTS:
            hit = gt_hits[gt]
            if hit is None or int(hit["length"]) < min_aligned_nt:
                row[f"GT{gt}_Distance"] = ""
                row[f"GT{gt}_AlignedNT"] = ""
            else:
                row[f"GT{gt}_Distance"] = hit["distance"]
                row[f"GT{gt}_AlignedNT"] = hit["length"]
        rows.append(row)
    return rows


def resistance_positions_covered(hit: dict[str, Any]) -> list[int]:
    sstart = min(int(hit["sstart"]), int(hit["send"]))
    send = max(int(hit["sstart"]), int(hit["send"]))
    covered: list[int] = []
    for pos in RESISTANCE_POSITIONS:
        pos_nt_start = ((pos - 1) * 3) + 1
        pos_nt_end = pos * 3
        if sstart <= pos_nt_end and send >= pos_nt_start:
            covered.append(pos)
    return covered


def write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "NS5B_GT_Distance"
    fieldnames = [
        "RefID",
        "RefName",
        "GenBankAccession",
        *[f"GT{gt}_Distance" for gt in REFERENCE_GTS],
        "BestGT",
        "BestGTDistance",
        "AlignedNT",
        "ContainsResistancePosition",
        "ResistancePositionsCovered",
        *[f"GT{gt}_AlignedNT" for gt in REFERENCE_GTS],
    ]
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])
    workbook.save(path)


# Historical helper kept for reference only. The current workflow writes only
# NS5B_GT_AllStudies.xlsx.
#
# def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
#     fieldnames = [
#         "RefID",
#         "RefName",
#         "GenBankAccession",
#         *[f"GT{gt}_Distance" for gt in REFERENCE_GTS],
#         "BestGT",
#         "BestGTDistance",
#         "AlignedNT",
#         *[f"GT{gt}_AlignedNT" for gt in REFERENCE_GTS],
#     ]
#     with path.open("w", encoding="utf-8", newline="") as handle:
#         writer = csv.DictWriter(handle, fieldnames=fieldnames)
#         writer.writeheader()
#         for row in rows:
#             writer.writerow(row)


def main() -> int:
    args = parse_args()
    excel_file = Path(args.excel_file).expanduser()
    fasta_dir = Path(args.fasta_dir).expanduser()
    reference_fasta = Path(args.reference_fasta).expanduser()
    base_output_dir = Path(args.output_dir)

    refs = load_reference_ns5b(reference_fasta)
    studies = load_filtered_studies(
        excel_file,
        args.sheet,
        args.refid_column,
        args.refname_column,
        args.numpatients_column,
        args.positive_column,
    )
    matched = find_study_fasta_files(fasta_dir, studies)

    job_dir = make_job_dir(base_output_dir, excel_file, args.sheet)
    output_path = base_output_dir / "NS5B_GT_AllStudies.xlsx"
    # Historical output flow kept for reference only.
    # progress_dir = job_dir / "NS5B_Alignments.xlsx"
    # progress_dir.mkdir(parents=True, exist_ok=True)
    db_prefix, subject_id_to_gt = build_reference_db(job_dir, refs)

    master_rows: list[dict[str, Any]] = []
    # Historical output flow kept for reference only.
    # study_summaries: list[dict[str, Any]] = []
    for study in matched:
        rows = build_rows_for_study(study, db_prefix, subject_id_to_gt, args.min_aligned_nt)
        # Historical per-study workbook flow kept for reference only.
        # safe_stem = sanitize_label(Path(study["fasta_path"]).stem)
        # xlsx_path = progress_dir / f"{safe_stem}.xlsx"
        # write_xlsx(xlsx_path, rows)
        master_rows.extend(rows)
        # study_summaries.append(
        #     {
        #         "RefID": study["RefID"],
        #         "RefName": study["RefName"],
        #         "fasta_file": Path(study["fasta_path"]).name,
        #         "sequence_rows_written": len(rows),
        #         "xlsx_file": str(xlsx_path.resolve()),
        #     }
        # )
    write_xlsx(output_path, master_rows)
    # Historical output flow kept for reference only.
    # write_csv(job_dir / "ns5b_gt_distance_master.csv", master_rows)
    # (job_dir / "study_progress.json").write_text(
    #     json.dumps(study_summaries, indent=2, ensure_ascii=True) + "\n",
    #     encoding="utf-8",
    # )
    # (job_dir / "workflow_request.txt").write_text(
    #     Path("notes/ns5b_gt_distance_workflow_2026-05-13.md").read_text(encoding="utf-8"),
    #     encoding="utf-8",
    # )

    payload = {
        "output_dir": str(base_output_dir.resolve()),
        "study_count": len(matched),
        "master_row_count": len(master_rows),
        "combined_xlsx": str(output_path.resolve()),
        # Historical payload fields kept for reference only.
        # "master_csv": str((job_dir / "ns5b_gt_distance_master.csv").resolve()),
        # "xlsx_dir": str(progress_dir.resolve()),
    }
    for suffix in (".nhr", ".nin", ".nsq", ".ndb", ".not", ".ntf", ".nto"):
        db_file = Path(f"{db_prefix}{suffix}")
        if db_file.exists():
            try:
                db_file.unlink()
            except OSError:
                pass
    try:
        (job_dir / "ns5b_gt_refs.fasta").unlink()
    except OSError:
        pass
    shutil.rmtree(job_dir, ignore_errors=True)
    print(json.dumps(payload, indent=2, ensure_ascii=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
