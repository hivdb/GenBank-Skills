#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


REFERENCE_AA_LENGTH = 213


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export NS5A consensus amino-acid FASTA records from GT and subtype complete-profile workbooks."
    )
    parser.add_argument("--gt-profile-workbook", required=True)
    parser.add_argument("--subtype-profile-workbook", required=True)
    parser.add_argument("--output-dir", default="outputs")
    parser.add_argument("--gt-output-fasta")
    parser.add_argument("--subtype-output-fasta")
    return parser.parse_args()


def script_temp_dir() -> Path:
    path = Path("temp") / "hcv-ns5a-build-workflow" / Path(__file__).stem
    path.mkdir(parents=True, exist_ok=True)
    return path


def infer_position_column(header: list[str]) -> str:
    position_columns = [name for name in header if name.endswith("Position") and name != "NumSeqsIncludingPosition"]
    if len(position_columns) != 1:
        raise RuntimeError(f"Expected exactly one position column, found: {position_columns}")
    return position_columns[0]


def load_gt_entries(workbook_path: Path) -> tuple[list[tuple[str, str]], dict[str, list[int]]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    entries: list[tuple[str, str]] = []
    positions_by_gt: dict[str, list[int]] = {}
    for sheet_name in wb.sheetnames:
        gt_label = sheet_name.strip()
        ws = wb[sheet_name]
        header = [str(value) if value is not None else "" for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        position_column = infer_position_column(header)
        index_by_name = {name: idx for idx, name in enumerate(header)}
        required = [position_column, "AminoAcid", "PctWithAA"]
        missing = [name for name in required if name not in index_by_name]
        if missing:
            raise RuntimeError(f"Worksheet {sheet_name} in {workbook_path} is missing columns: {', '.join(missing)}")

        best_by_position: dict[int, tuple[float, str]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            position = int(row[index_by_name[position_column]])
            if position < 1 or position > REFERENCE_AA_LENGTH:
                continue
            amino_acid = str(row[index_by_name["AminoAcid"]]).strip()
            pct_with_aa = float(row[index_by_name["PctWithAA"]])
            current = best_by_position.get(position)
            candidate = (pct_with_aa, amino_acid)
            if current is None or candidate[0] > current[0] or (candidate[0] == current[0] and candidate[1] < current[1]):
                best_by_position[position] = candidate

        ordered_positions = sorted(best_by_position)
        if not ordered_positions:
            raise RuntimeError(f"Worksheet {sheet_name} in {workbook_path} has no position rows")
        sequence = "".join(best_by_position[position][1] for position in ordered_positions)
        entries.append((gt_label, sequence))
        positions_by_gt[gt_label] = ordered_positions
    wb.close()
    if not entries:
        raise RuntimeError(f"No GT blocks found in {workbook_path}")
    return entries, positions_by_gt


def load_subtype_entries(workbook_path: Path) -> tuple[list[tuple[str, str]], dict[str, list[int]]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    entries: list[tuple[str, str]] = []
    positions_by_subtype: dict[str, list[int]] = {}
    for sheet_name in wb.sheetnames:
        gt_label = sheet_name.strip()
        ws = wb[sheet_name]
        header = [str(value) if value is not None else "" for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        position_column = infer_position_column(header)
        index_by_name = {name: idx for idx, name in enumerate(header)}
        required = ["Subtype", position_column, "AminoAcid", "PctWithAA"]
        missing = [name for name in required if name not in index_by_name]
        if missing:
            raise RuntimeError(f"Worksheet {sheet_name} in {workbook_path} is missing columns: {', '.join(missing)}")

        best_by_subtype_position: dict[str, dict[int, tuple[float, str]]] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            subtype = str(row[index_by_name["Subtype"]]).strip()
            position = int(row[index_by_name[position_column]])
            if position < 1 or position > REFERENCE_AA_LENGTH:
                continue
            amino_acid = str(row[index_by_name["AminoAcid"]]).strip()
            pct_with_aa = float(row[index_by_name["PctWithAA"]])
            per_position = best_by_subtype_position.setdefault(subtype, {})
            current = per_position.get(position)
            candidate = (pct_with_aa, amino_acid)
            if current is None or candidate[0] > current[0] or (candidate[0] == current[0] and candidate[1] < current[1]):
                per_position[position] = candidate

        for subtype in sorted(best_by_subtype_position):
            ordered_positions = sorted(best_by_subtype_position[subtype])
            if not ordered_positions:
                continue
            sequence = "".join(best_by_subtype_position[subtype][position][1] for position in ordered_positions)
            header_label = f"{gt_label}_{subtype}"
            entries.append((header_label, sequence))
            positions_by_subtype[header_label] = ordered_positions
    wb.close()
    if not entries:
        raise RuntimeError(f"No subtype rows found in {workbook_path}")
    return entries, positions_by_subtype


def write_fasta(path: Path, entries: list[tuple[str, str]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for header, sequence in entries:
            handle.write(f">{header}\n")
            for start in range(0, len(sequence), 70):
                handle.write(sequence[start : start + 70] + "\n")


def main() -> int:
    args = parse_args()
    script_temp_dir()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    gt_workbook = Path(args.gt_profile_workbook).expanduser()
    subtype_workbook = Path(args.subtype_profile_workbook).expanduser()
    gt_fasta = Path(args.gt_output_fasta).expanduser() if args.gt_output_fasta else output_dir / "NS5A_GT_Consensus.fasta"
    subtype_fasta = (
        Path(args.subtype_output_fasta).expanduser()
        if args.subtype_output_fasta
        else output_dir / "NS5A_Subtype_Consensus.fasta"
    )

    gt_entries, positions_by_gt = load_gt_entries(gt_workbook)
    subtype_entries, positions_by_subtype = load_subtype_entries(subtype_workbook)
    write_fasta(gt_fasta, gt_entries)
    write_fasta(subtype_fasta, subtype_entries)

    summary = {
        "gene": "NS5A",
        "gt_fasta": str(gt_fasta.resolve()),
        "gt_entry_count": len(gt_entries),
        "positions_by_gt": positions_by_gt,
        "subtype_fasta": str(subtype_fasta.resolve()),
        "subtype_entry_count": len(subtype_entries),
        "positions_by_subtype": positions_by_subtype,
    }
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
