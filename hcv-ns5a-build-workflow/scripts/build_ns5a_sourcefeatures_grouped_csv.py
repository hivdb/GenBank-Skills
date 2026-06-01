#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import re
import tempfile
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook


DEFAULT_GT_WORKBOOK = Path("outputs/NS5A_GT_AllStudies.xlsx")
DEFAULT_SUMMARY_XLSX = Path("outputs/NS5A_NumSeqs_Naive_1PP_CoversRAS_ByStudy.xlsx")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Group NS5A source-feature rows by RefID, then regroup by source_clone "
            "or source_isolate normalization rules and write one row per final group."
        )
    )
    parser.add_argument("--input-csv", default="", help="Path to NS5A sourcefeatures CSV")
    parser.add_argument("--output-csv", default="", help="Output grouped CSV path")
    parser.add_argument("--gt-workbook", default=str(DEFAULT_GT_WORKBOOK), help="Path to NS5A_GT_AllStudies.xlsx")
    parser.add_argument("--summary-xlsx", default=str(DEFAULT_SUMMARY_XLSX), help="Output Excel summary path")
    return parser.parse_args()


def script_temp_dir() -> Path:
    path = Path("temp") / Path(__file__).stem
    path.mkdir(parents=True, exist_ok=True)
    return path


def sourcefeatures_temp_csv_path() -> Path:
    return Path("temp") / "build_ns5a_sourcefeatures_csv" / "NS5A_SourceFeatures.csv"


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip())


def normalize_clone_key(value: str) -> str:
    text = normalize_space(value)
    if not text:
        return ""
    parts = [part for part in text.split("_") if part]
    if len(parts) > 1:
        return "_".join(parts[:-1]).strip("_")
    return text


def normalize_isolate_key(value: str) -> str:
    text = normalize_space(value)
    if not text:
        return ""
    match = re.search(r"(?i)\b(week|day)\b", text)
    if match:
        text = text[: match.start()]
    return text.strip(" _-")


def patient_code_from_structured_comment(value: str) -> str:
    match = re.search(r"(?im)^Patient Code\s*::\s*(.+?)\s*$", value)
    return normalize_space(match.group(1)) if match else ""


def effective_source_isolate(row: dict[str, str]) -> str:
    source_isolate = (row.get("source_isolate") or "").strip()
    if not source_isolate:
        return ""
    patient_code = patient_code_from_structured_comment((row.get("StructuredComment") or "").strip())
    if patient_code and patient_code in source_isolate:
        return patient_code
    return source_isolate


def read_rows(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = [dict(row) for row in reader]
        fieldnames = reader.fieldnames or []
    return rows, fieldnames


def load_gt_coverage_by_accession(path: Path) -> dict[str, list[int]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    header = [str(value) if value is not None else "" for value in next(sheet.iter_rows(values_only=True))]
    index = {name: idx for idx, name in enumerate(header)}
    required = ["GenBankAccession", "ResistancePositionsCovered"]
    for name in required:
        if name not in index:
            raise RuntimeError(f"Column '{name}' was not found in {path}")
    coverage: dict[str, list[int]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        accession = str(row[index["GenBankAccession"]] or "").strip()
        if not accession:
            continue
        covered_text = str(row[index["ResistancePositionsCovered"]] or "").strip()
        positions = [int(part) for part in covered_text.split(",") if part.strip()]
        coverage[accession] = positions
    workbook.close()
    return coverage


def choose_best_coverage_accession(rows: list[dict[str, str]], gt_coverage_by_accession: dict[str, list[int]]) -> tuple[str, list[int]]:
    best_accession = ""
    best_positions: list[int] = []
    for row in rows:
        accession = (row.get("Accession") or "").strip()
        positions = gt_coverage_by_accession.get(accession, [])
        candidate = (len(positions), positions, accession)
        current = (len(best_positions), best_positions, best_accession)
        if candidate > current:
            best_accession = accession
            best_positions = positions
    return best_accession, best_positions


def summarize_group(refid: str, mode: str, group_key: str, rows: list[dict[str, str]], gt_coverage_by_accession: dict[str, list[int]]) -> dict[str, str]:
    row_count = len(rows)
    accessions = sorted({(row.get("Accession") or "").strip() for row in rows if (row.get("Accession") or "").strip()})
    source_clones = sorted({(row.get("source_clone") or "").strip() for row in rows if (row.get("source_clone") or "").strip()})
    source_isolates = sorted({effective_source_isolate(row) for row in rows if effective_source_isolate(row)})
    definitions = sorted({(row.get("definition") or "").strip() for row in rows if (row.get("definition") or "").strip()})
    structured_comments = sorted({(row.get("StructuredComment") or "").strip() for row in rows if (row.get("StructuredComment") or "").strip()})
    best_accession, best_positions = choose_best_coverage_accession(rows, gt_coverage_by_accession)
    return {
        "RefID": refid,
        "GroupingMode": "" if row_count == 1 else mode,
        "GroupKey": group_key,
        "RowCount": str(row_count),
        "Accessions": " | ".join(accessions),
        "SelectedAccessionMostResistanceCovered": best_accession,
        "CoveredResistancePositions": ",".join(str(pos) for pos in best_positions),
        "CoveredResistancePositionCount": str(len(best_positions)),
        "source_clone_values": " | ".join(source_clones),
        "source_isolate_values": " | ".join(source_isolates),
        "definition_values": " | ".join(definitions),
        "StructuredComment_values": " | ".join(structured_comments),
    }


def group_rows(rows: list[dict[str, str]], gt_coverage_by_accession: dict[str, list[int]]) -> list[dict[str, str]]:
    by_refid: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        by_refid[(row.get("RefID") or "").strip()].append(row)
    grouped_rows: list[dict[str, str]] = []
    for refid in sorted(by_refid):
        ref_rows = by_refid[refid]
        clone_present: list[dict[str, str]] = []
        clone_empty: list[dict[str, str]] = []
        for row in ref_rows:
            if (row.get("source_clone") or "").strip():
                clone_present.append(row)
            else:
                clone_empty.append(row)
        clone_groups: dict[str, list[dict[str, str]]] = defaultdict(list)
        for row in clone_present:
            clone_groups[normalize_clone_key((row.get("source_clone") or "").strip())].append(row)
        for key in sorted(clone_groups):
            grouped_rows.append(summarize_group(refid, "source_clone_prefix", key, clone_groups[key], gt_coverage_by_accession))
        isolate_groups: dict[str, list[dict[str, str]]] = defaultdict(list)
        for row in clone_empty:
            raw_isolate = effective_source_isolate(row)
            key = normalize_isolate_key(raw_isolate)
            if not key:
                accession = (row.get("Accession") or "").strip()
                key = f"__ungrouped__{accession}"
            isolate_groups[key].append(row)
        for key in sorted(isolate_groups):
            grouped_rows.append(summarize_group(refid, "source_isolate_multiple_timepoints", key, isolate_groups[key], gt_coverage_by_accession))
    return grouped_rows


def write_rows(path: Path, rows: list[dict[str, str]]) -> None:
    fieldnames = [
        "RefID",
        "GroupingMode",
        "GroupKey",
        "RowCount",
        "Accessions",
        "SelectedAccessionMostResistanceCovered",
        "CoveredResistancePositions",
        "CoveredResistancePositionCount",
        "source_clone_values",
        "source_isolate_values",
        "definition_values",
        "StructuredComment_values",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_refid_summary_rows(rows: list[dict[str, str]]) -> list[dict[str, int | str]]:
    counts_by_refid: dict[str, int] = defaultdict(int)
    for row in rows:
        if int(row.get("CoveredResistancePositionCount") or 0) <= 0:
            continue
        counts_by_refid[row.get("RefID", "")] += 1
    return [{"RefID": refid, "1PP_CoversRAS": count} for refid, count in sorted(counts_by_refid.items())]


def write_summary_xlsx(path: Path, rows: list[dict[str, int | str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "NS5A_1PP_CoversRAS"
    fieldnames = ["RefID", "1PP_CoversRAS"]
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])
    workbook.save(path)


def main() -> int:
    args = parse_args()
    input_csv = Path(args.input_csv).expanduser() if args.input_csv else sourcefeatures_temp_csv_path()
    output_csv = Path(args.output_csv).expanduser() if args.output_csv else script_temp_dir() / "NS5A_SourceFeatures_Grouped.csv"
    gt_workbook = Path(args.gt_workbook).expanduser()
    summary_xlsx = Path(args.summary_xlsx).expanduser()
    if not input_csv.is_file():
        raise RuntimeError(f"Input CSV was not found: {input_csv}")
    if not gt_workbook.is_file():
        raise RuntimeError(f"GT workbook was not found: {gt_workbook}")
    rows, _ = read_rows(input_csv)
    gt_coverage_by_accession = load_gt_coverage_by_accession(gt_workbook)
    grouped_rows = group_rows(rows, gt_coverage_by_accession)
    write_rows(output_csv, grouped_rows)
    summary_rows = build_refid_summary_rows(grouped_rows)
    write_summary_xlsx(summary_xlsx, summary_rows)
    print(
        {
            "input_csv": str(input_csv.resolve()),
            "gt_workbook": str(gt_workbook.resolve()),
            "output_csv": str(output_csv.resolve()),
            "summary_xlsx": str(summary_xlsx.resolve()),
            "input_row_count": len(rows),
            "output_row_count": len(grouped_rows),
            "summary_row_count": len(summary_rows),
        }
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
