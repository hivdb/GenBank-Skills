from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Add a genotype-count summary sheet to an NS3 combined workbook."
    )
    parser.add_argument(
        "--workbook",
        required=True,
        type=Path,
        help="Path to the combined Excel workbook.",
    )
    parser.add_argument(
        "--data-sheet",
        default=None,
        help="Worksheet containing the sequence-level rows. Defaults to the first sheet.",
    )
    parser.add_argument(
        "--summary-sheet",
        default="Genotype_Counts",
        help="Name of the summary worksheet to create or replace.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook

    wb = load_workbook(workbook_path)
    ws = wb[args.data_sheet] if args.data_sheet else wb[wb.sheetnames[0]]

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        best_gt_idx = header.index("BestGT") + 1
    except ValueError as exc:
        raise RuntimeError("Column 'BestGT' was not found in the workbook.") from exc

    counts: Counter[str] = Counter()
    total_rows = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        best_gt = row[best_gt_idx - 1]
        if best_gt is None or str(best_gt).strip() == "":
            continue
        counts[str(best_gt).strip()] += 1
        total_rows += 1

    if args.summary_sheet in wb.sheetnames:
        del wb[args.summary_sheet]
    summary_ws = wb.create_sheet(args.summary_sheet)
    summary_ws.append(["BestGT", "SequenceCount"])
    for gt in sorted(counts, key=lambda value: (int(value) if value.isdigit() else value)):
        summary_ws.append([gt, counts[gt]])
    summary_ws.append(["Total", total_rows])

    wb.save(workbook_path)

    print(f"updated_workbook={workbook_path}")
    print(f"summary_sheet={args.summary_sheet}")
    for gt in sorted(counts, key=lambda value: (int(value) if value.isdigit() else value)):
        print(f"GT{gt}={counts[gt]}")
    print(f"Total={total_rows}")


if __name__ == "__main__":
    main()
