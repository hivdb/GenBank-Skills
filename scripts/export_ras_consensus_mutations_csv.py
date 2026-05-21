#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path

from openpyxl import load_workbook


RAS_POSITIONS = {
    "NS3": [36, 41, 43, 54, 55, 56, 80, 122, 155, 156, 158, 166, 168, 170, 175],
    "NS5A": [24, 26, 28, 29, 30, 31, 32, 38, 58, 62, 92, 93],
    "NS5B": [150, 159, 206, 282, 316, 320, 321],
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export detailed and aggregated RAS consensus CSVs from GT and subtype complete-profile workbooks."
    )
    parser.add_argument("--gene", required=True, choices=["NS3", "NS5A", "NS5B"])
    parser.add_argument("--gt-profile-workbook", required=True)
    parser.add_argument("--subtype-profile-workbook", required=True)
    parser.add_argument("--output-dir", default="outputs")
    return parser.parse_args()


def script_temp_dir() -> Path:
    path = Path("temp") / Path(__file__).stem
    path.mkdir(parents=True, exist_ok=True)
    return path


def infer_position_column(header: list[str]) -> str:
    position_columns = [name for name in header if name.endswith("Position") and name != "NumSeqsIncludingPosition"]
    if len(position_columns) != 1:
        raise RuntimeError(f"Expected exactly one position column, found: {position_columns}")
    return position_columns[0]


def genotype_sort_key(genotype: str) -> tuple[int, str]:
    text = genotype.strip()
    if text.startswith("GT"):
        text = text[2:]
    try:
        return (int(text), genotype)
    except ValueError:
        return (999, genotype)


def subtype_sort_key(subtype: str) -> tuple[int, str]:
    return (0 if subtype == "" else 1, subtype)


def load_gt_rows(workbook_path: Path, gene: str) -> list[dict[str, object]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    rows_out: list[dict[str, object]] = []
    wanted = set(RAS_POSITIONS[gene])
    for sheet_name in wb.sheetnames:
        genotype = sheet_name.strip()
        ws = wb[sheet_name]
        header = [str(value) if value is not None else "" for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        position_column = infer_position_column(header)
        index_by_name = {name: idx for idx, name in enumerate(header)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            position = int(row[index_by_name[position_column]])
            pct_with_aa = float(row[index_by_name["PctWithAA"]])
            if position not in wanted or pct_with_aa < 10.0:
                continue
            rows_out.append(
                {
                    "Genotype": genotype,
                    "Subtype": "",
                    "Position": position,
                    position_column: position,
                    "NumSeqsIncludingPosition": int(row[index_by_name["NumSeqsIncludingPosition"]]),
                    "AminoAcid": str(row[index_by_name["AminoAcid"]]).strip(),
                    "CountWithAA": int(row[index_by_name["CountWithAA"]]),
                    "CountWithAAAlone": int(row[index_by_name["CountWithAAAlone"]]),
                    "PctWithAA": pct_with_aa,
                    "PctWithAAAlone": float(row[index_by_name["PctWithAAAlone"]]),
                }
            )
    wb.close()
    return rows_out


def load_subtype_rows(workbook_path: Path, gene: str) -> list[dict[str, object]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    rows_out: list[dict[str, object]] = []
    wanted = set(RAS_POSITIONS[gene])
    for sheet_name in wb.sheetnames:
        genotype = sheet_name.strip()
        ws = wb[sheet_name]
        header = [str(value) if value is not None else "" for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        position_column = infer_position_column(header)
        index_by_name = {name: idx for idx, name in enumerate(header)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            position = int(row[index_by_name[position_column]])
            pct_with_aa = float(row[index_by_name["PctWithAA"]])
            if position not in wanted or pct_with_aa < 10.0:
                continue
            rows_out.append(
                {
                    "Genotype": genotype,
                    "Subtype": str(row[index_by_name["Subtype"]]).strip(),
                    "Position": position,
                    position_column: position,
                    "NumSeqsIncludingPosition": int(row[index_by_name["NumSeqsIncludingPosition"]]),
                    "AminoAcid": str(row[index_by_name["AminoAcid"]]).strip(),
                    "CountWithAA": int(row[index_by_name["CountWithAA"]]),
                    "CountWithAAAlone": int(row[index_by_name["CountWithAAAlone"]]),
                    "PctWithAA": pct_with_aa,
                    "PctWithAAAlone": float(row[index_by_name["PctWithAAAlone"]]),
                }
            )
    wb.close()
    return rows_out


def write_detailed_csv(path: Path, rows: list[dict[str, object]], gene: str) -> None:
    position_column = f"{gene}Position"
    fieldnames = [
        "Genotype",
        "Subtype",
        position_column,
        "NumSeqsIncludingPosition",
        "AminoAcid",
        "CountWithAA",
        "CountWithAAAlone",
        "PctWithAA",
        "PctWithAAAlone",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def build_aggregate_rows(rows: list[dict[str, object]]) -> list[dict[str, str | int]]:
    grouped: dict[tuple[str, str, int], list[dict[str, object]]] = {}
    for row in rows:
        key = (str(row["Genotype"]), str(row["Subtype"]), int(row["Position"]))
        grouped.setdefault(key, []).append(row)

    result: list[dict[str, str | int]] = []
    for key in sorted(grouped, key=lambda item: (genotype_sort_key(item[0]), subtype_sort_key(item[1]), item[2])):
        group_rows = grouped[key]
        ordered = sorted(
            group_rows,
            key=lambda row: (-float(row["PctWithAA"]), str(row["AminoAcid"])),
        )
        result.append(
            {
                "Genotype": key[0],
                "Subtype": key[1],
                "Position": key[2],
                "AminoAcids": "".join(str(row["AminoAcid"]) for row in ordered),
            }
        )
    return result


def write_aggregate_csv(path: Path, rows: list[dict[str, str | int]]) -> None:
    fieldnames = ["Genotype", "Subtype", "Position", "AminoAcids"]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def main() -> int:
    args = parse_args()
    temp_dir = script_temp_dir()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    gt_rows = load_gt_rows(Path(args.gt_profile_workbook).expanduser(), args.gene)
    subtype_rows = load_subtype_rows(Path(args.subtype_profile_workbook).expanduser(), args.gene)
    all_rows = sorted(
        gt_rows + subtype_rows,
        key=lambda row: (
            genotype_sort_key(str(row["Genotype"])),
            subtype_sort_key(str(row["Subtype"])),
            int(row["Position"]),
            -float(row["PctWithAA"]),
            str(row["AminoAcid"]),
        ),
    )

    detailed_path = temp_dir / f"{args.gene}_RAS_Consensus_Mutations.csv"
    aggregate_path = output_dir / f"{args.gene}_RAS_Consensus.csv"

    write_detailed_csv(detailed_path, all_rows, args.gene)
    aggregate_rows = build_aggregate_rows(all_rows)
    write_aggregate_csv(aggregate_path, aggregate_rows)

    print(
        json.dumps(
            {
                "gene": args.gene,
                "detailed_csv": str(detailed_path.resolve()),
                "aggregate_csv": str(aggregate_path.resolve()),
                "detailed_row_count": len(all_rows),
                "aggregate_row_count": len(aggregate_rows),
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
