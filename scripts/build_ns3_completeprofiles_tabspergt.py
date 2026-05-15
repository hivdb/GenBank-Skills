#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import shutil
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


AA_ORDER = list("ACDEFGHIKLMNPQRSTVWY") + ["X", "*"]
SUBTYPE_MIN_PCT = 10.0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build GT and subtype NS3 amino-acid profile workbooks.")
    parser.add_argument("--input-workbook", required=True, help="Path to NS3 AA extraction workbook.")
    parser.add_argument("--output-dir", default="outputs", help="Base output directory.")
    return parser.parse_args()


def script_temp_dir() -> Path:
    path = Path("temp") / Path(__file__).stem
    path.mkdir(parents=True, exist_ok=True)
    return path


def sanitize_label(value: str) -> str:
    return "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in value).strip("._-") or "job"


def make_job_dir(base_output_dir: Path, workbook_path: Path) -> Path:
    label = sanitize_label(f"{workbook_path.stem}_ns3_aa_profiles")
    job_dir = base_output_dir / label
    if job_dir.exists():
        shutil.rmtree(job_dir)
    job_dir.mkdir(parents=True, exist_ok=True)
    return job_dir


def load_rows(workbook_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(v) if v is not None else "" for v in next(ws.iter_rows(values_only=True))]
    index = {name: i for i, name in enumerate(header)}
    required = ["ClosestGT", "ClosestSubtype", "StartAAPosition", "EndAAPosition", "AASequence"]
    for name in required:
        if name not in index:
            raise RuntimeError(f"Column '{name}' not found in {workbook_path}")
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        aa_sequence = values[index["AASequence"]]
        start = values[index["StartAAPosition"]]
        end = values[index["EndAAPosition"]]
        if not aa_sequence or start in (None, "") or end in (None, ""):
            continue
        rows.append(
            {
                "ClosestGT": str(values[index["ClosestGT"]]).strip(),
                "ClosestSubtype": str(values[index["ClosestSubtype"]]).strip(),
                "StartAAPosition": int(start),
                "EndAAPosition": int(end),
                "AASequence": str(aa_sequence).strip(),
            }
        )
    wb.close()
    return rows


def build_position_counts(rows: list[dict[str, Any]]) -> tuple[dict[int, int], dict[int, Counter[str]]]:
    included_counts: dict[int, int] = defaultdict(int)
    aa_counts: dict[int, Counter[str]] = defaultdict(Counter)
    for row in rows:
        start = row["StartAAPosition"]
        aa_sequence = row["AASequence"]
        for offset, aa in enumerate(aa_sequence):
            pos = start + offset
            included_counts[pos] += 1
            aa_counts[pos][aa] += 1
    return included_counts, aa_counts


def write_gt_workbook(path: Path, rows_by_gt: dict[str, list[dict[str, Any]]]) -> dict[str, int]:
    wb = Workbook()
    wb.remove(wb.active)
    summary: dict[str, int] = {}
    header = [
        "NS3Position",
        "NumSeqsIncludingPosition",
        "AminoAcid",
        "CountWithAA",
        "CountWithAAAlone",
        "PctWithAA",
        "PctWithAAAlone",
    ]
    for gt in sorted(rows_by_gt, key=int):
        ws = wb.create_sheet(f"GT{gt}")
        ws.append(header)
        included_counts, aa_counts = build_position_counts(rows_by_gt[gt])
        summary[gt] = len(rows_by_gt[gt])
        for pos in sorted(included_counts):
            denom = included_counts[pos]
            for aa in AA_ORDER:
                count = aa_counts[pos].get(aa, 0)
                if count == 0:
                    continue
                ws.append([pos, denom, aa, count, count, 100.0 * count / denom, 100.0 * count / denom])
    wb.save(path)
    return summary


def write_subtype_workbook(path: Path, rows_by_gt_subtype: dict[str, dict[str, list[dict[str, Any]]]]) -> dict[str, dict[str, int]]:
    wb = Workbook()
    wb.remove(wb.active)
    summary: dict[str, dict[str, int]] = {}
    header = [
        "Subtype",
        "NS3Position",
        "NumSeqsIncludingPosition",
        "AminoAcid",
        "CountWithAA",
        "CountWithAAAlone",
        "PctWithAA",
        "PctWithAAAlone",
    ]
    for gt in sorted(rows_by_gt_subtype, key=int):
        ws = wb.create_sheet(f"GT{gt}")
        ws.append(header)
        summary[gt] = {}
        for subtype in sorted(rows_by_gt_subtype[gt]):
            subtype_rows = rows_by_gt_subtype[gt][subtype]
            included_counts, aa_counts = build_position_counts(subtype_rows)
            summary[gt][subtype] = len(subtype_rows)
            for pos in sorted(included_counts):
                denom = included_counts[pos]
                for aa in AA_ORDER:
                    count = aa_counts[pos].get(aa, 0)
                    if count == 0:
                        continue
                    pct = 100.0 * count / denom
                    if pct < SUBTYPE_MIN_PCT:
                        continue
                    ws.append([subtype, pos, denom, aa, count, count, pct, pct])
    wb.save(path)
    return summary


def main() -> int:
    args = parse_args()
    input_workbook = Path(args.input_workbook).expanduser()
    output_dir = Path(args.output_dir)
    script_temp_dir()

    rows = load_rows(input_workbook)
    rows_by_gt: dict[str, list[dict[str, Any]]] = defaultdict(list)
    rows_by_gt_subtype: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
    for row in rows:
        gt = row["ClosestGT"]
        subtype = row["ClosestSubtype"]
        rows_by_gt[gt].append(row)
        rows_by_gt_subtype[gt][subtype].append(row)

    output_dir.mkdir(parents=True, exist_ok=True)
    gt_path = output_dir / "NS3_GT_CompleteProfiles_TabsPerGT.xlsx"
    subtype_path = output_dir / "NS3_Subtype_CompleteProfiles_TabsPerGT.xlsx"

    gt_summary = write_gt_workbook(gt_path, rows_by_gt)
    subtype_summary = write_subtype_workbook(subtype_path, rows_by_gt_subtype)

    summary = {
        "input_workbook": str(input_workbook.resolve()),
        "rows_with_aa": len(rows),
        "gt_workbook": str(gt_path.resolve()),
        "subtype_workbook": str(subtype_path.resolve()),
        "subtype_min_percent": SUBTYPE_MIN_PCT,
        "gt_sequence_counts": gt_summary,
        "subtype_group_count": sum(len(v) for v in rows_by_gt_subtype.values()),
        "note": "CountWithAA and CountWithAAAlone are identical because current AA sequences contain single-letter calls, not explicit mixtures.",
    }
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
