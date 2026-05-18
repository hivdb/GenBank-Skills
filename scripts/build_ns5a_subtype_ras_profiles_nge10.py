#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_RESISTANCE_POSITIONS = [24, 26, 28, 29, 30, 31, 32, 38, 58, 62, 92, 93]
EXCLUDED_AAS = {"X", "*"}
TARGET_GENE = "NS5A"
AA_REF_GENE = "NS5A_NTD"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build subtype-level NS5A resistance-position AA profile summary in Excel."
    )
    parser.add_argument("--subtype-profile-workbook", required=True)
    parser.add_argument("--gt-aa-json", required=True)
    parser.add_argument("--output-dir", default="outputs")
    parser.add_argument("--min-sequences", type=int, default=10)
    parser.add_argument(
        "--positions",
        default=",".join(str(pos) for pos in DEFAULT_RESISTANCE_POSITIONS),
        help="Comma-separated amino-acid positions to summarize.",
    )
    return parser.parse_args()


def script_temp_dir() -> Path:
    path = Path("temp") / Path(__file__).stem
    path.mkdir(parents=True, exist_ok=True)
    return path


def parse_positions(raw: str) -> list[int]:
    positions: list[int] = []
    for token in raw.split(","):
        text = token.strip()
        if not text:
            continue
        positions.append(int(text))
    if not positions:
        raise RuntimeError("At least one resistance position is required.")
    return positions


def sanitize_label(value: str) -> str:
    return "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in value).strip("._-") or "job"


def make_job_dir(base_output_dir: Path, workbook_path: Path) -> Path:
    label = sanitize_label(f"{workbook_path.stem}_ns5a_subtype_resistance_profile")
    job_dir = base_output_dir / label
    if job_dir.exists():
        shutil.rmtree(job_dir)
    job_dir.mkdir(parents=True, exist_ok=True)
    return job_dir


def format_freq(value: float) -> str:
    if value <= 0:
        return "0"
    sig = 2 if value >= 1.0 else 1
    decimals = max(0, sig - 1 - math.floor(math.log10(value)))
    text = f"{round(value, decimals):.{decimals}f}"
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text


def load_consensus_by_gt(json_path: Path) -> dict[str, str]:
    rows = json.loads(json_path.read_text(encoding="utf-8"))
    consensus: dict[str, str] = {}
    for row in rows:
        name = str(row.get("name", ""))
        match = re.fullmatch(r"HCV([1-8])NS5A_NTD", name)
        if match:
            consensus[match.group(1)] = str(row.get("refSequence", "")).strip().upper()
    return consensus


def load_subtype_profile_rows(
    workbook_path: Path,
    positions: list[int],
) -> tuple[
    dict[str, dict[str, dict[int, list[tuple[str, float]]]]],
    dict[str, dict[str, int]],
    dict[str, dict[str, dict[int, int]]],
]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    profile_rows: dict[str, dict[str, dict[int, list[tuple[str, float]]]]] = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    subtype_counts: dict[str, dict[str, int]] = defaultdict(dict)
    position_coverage: dict[str, dict[str, dict[int, int]]] = defaultdict(lambda: defaultdict(dict))
    wanted = set(positions)
    for sheet_name in wb.sheetnames:
        gt = sheet_name.replace("GT", "")
        ws = wb[sheet_name]
        next(ws.iter_rows(values_only=True))
        for row in ws.iter_rows(min_row=2, values_only=True):
            subtype = str(row[0])
            pos = int(row[1])
            denom = int(row[2])
            aa = str(row[3])
            pct = float(row[6])
            if pos in wanted:
                profile_rows[gt][subtype][pos].append((aa, pct))
                position_coverage[gt][subtype][pos] = denom
            current = subtype_counts[gt].get(subtype, 0)
            if denom > current:
                subtype_counts[gt][subtype] = denom
    wb.close()
    return profile_rows, subtype_counts, position_coverage


def build_grid(
    consensus_by_gt: dict[str, str],
    profile_rows: dict[str, dict[str, dict[int, list[tuple[str, float]]]]],
    subtype_counts: dict[str, dict[str, int]],
    position_coverage: dict[str, dict[str, dict[int, int]]],
    min_sequences: int,
    positions: list[int],
) -> list[list[str]]:
    grid: list[list[str]] = []
    ordered_subtypes: list[tuple[str, str]] = []
    for gt in sorted(subtype_counts, key=int):
        for subtype in sorted(subtype_counts[gt]):
            if subtype_counts[gt][subtype] >= min_sequences:
                ordered_subtypes.append((gt, subtype))

    for gt, subtype in ordered_subtypes:
        consensus_seq = consensus_by_gt[gt]
        pos_variants: dict[int, list[str]] = {}
        max_depth = 0
        for pos in positions:
            variants = [
                f"{aa}-{format_freq(pct)}"
                for aa, pct in sorted(profile_rows[gt][subtype].get(pos, []), key=lambda item: (-item[1], item[0]))
                if aa not in EXCLUDED_AAS and pct > 0.1
            ]
            pos_variants[pos] = variants
            max_depth = max(max_depth, len(variants))

        grid.append([f"GT{gt}_{subtype} ({subtype_counts[gt][subtype]})"] + [""] * len(positions))
        grid.append(["Position"] + [str(pos) for pos in positions])
        grid.append(["Reference"] + [consensus_seq[pos - 1] for pos in positions])
        coverage_row = ["Coverage"]
        total_sequences = subtype_counts[gt][subtype]
        for pos in positions:
            covered = position_coverage[gt][subtype].get(pos, 0)
            pct = (100.0 * covered / total_sequences) if total_sequences else 0.0
            coverage_row.append(f"{covered}/{total_sequences} ({pct:.1f}%)")
        grid.append(coverage_row)
        for depth in range(max_depth):
            row = ["Consensus" if depth == 0 else f"Rank{depth + 1}"]
            for pos in positions:
                variants = pos_variants[pos]
                row.append(variants[depth] if depth < len(variants) else "")
            grid.append(row)
        grid.append([""] + [""] * len(positions))
    return grid


def write_excel(path: Path, grid: list[list[str]], positions: list[int]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Subtype_Resistance_Profile"
    block_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    consensus_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    bold = Font(bold=True)

    for row_idx, row in enumerate(grid, start=1):
        ws.append(row)
        first = row[0]
        if first.startswith("GT"):
            for cell in ws[row_idx]:
                cell.fill = block_fill
                cell.font = bold
        elif first == "Position":
            for cell in ws[row_idx]:
                cell.fill = header_fill
                cell.font = bold
        elif first == "Reference":
            for cell in ws[row_idx]:
                cell.fill = consensus_fill
                cell.font = bold
        for cell in ws[row_idx]:
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(positions) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12 if col > 1 else 18
    wb.save(path)


def main() -> int:
    args = parse_args()
    subtype_profile_workbook = Path(args.subtype_profile_workbook).expanduser()
    gt_aa_json = Path(args.gt_aa_json).expanduser()
    output_dir = Path(args.output_dir)
    positions = parse_positions(args.positions)
    script_temp_dir()

    consensus_by_gt = load_consensus_by_gt(gt_aa_json)
    profile_rows, subtype_counts, position_coverage = load_subtype_profile_rows(subtype_profile_workbook, positions)
    grid = build_grid(consensus_by_gt, profile_rows, subtype_counts, position_coverage, args.min_sequences, positions)

    output_dir.mkdir(parents=True, exist_ok=True)
    excel_path = output_dir / "NS5A_Subtype_RAS_Profiles_NGE10.xlsx"
    write_excel(excel_path, grid, positions)

    included = []
    for gt in sorted(subtype_counts, key=int):
        for subtype in sorted(subtype_counts[gt]):
            if subtype_counts[gt][subtype] >= args.min_sequences:
                included.append({"gt": gt, "subtype": subtype, "count": subtype_counts[gt][subtype]})

    summary = {
        "excel": str(excel_path.resolve()),
        "gene": TARGET_GENE,
        "positions": positions,
        "frequency_threshold_percent": 0.1,
        "min_sequences": args.min_sequences,
        "included_subtypes": included,
    }
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
