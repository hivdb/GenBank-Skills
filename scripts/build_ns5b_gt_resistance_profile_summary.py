#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


DEFAULT_RESISTANCE_POSITIONS = [159, 282, 316, 320, 321, 414, 446, 553, 554, 556, 559, 561]
EXCLUDED_AAS = {"X", "*"}
TARGET_GENE = "NS5B"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build GT-level NS5B resistance-position AA profile summary in Excel and PNG."
    )
    parser.add_argument("--gt-profile-workbook", required=True)
    parser.add_argument("--gt-aa-json", required=True)
    parser.add_argument("--output-dir", default="outputs")
    parser.add_argument(
        "--positions",
        default=",".join(str(pos) for pos in DEFAULT_RESISTANCE_POSITIONS),
        help="Comma-separated amino-acid positions to summarize.",
    )
    return parser.parse_args()


def parse_positions(raw: str) -> list[int]:
    positions: list[int] = []
    for token in raw.split(","):
        text = token.strip()
        if not text:
            continue
        positions.append(int(text))
    if not positions:
        raise RuntimeError("At least one resistance position is required.")
    return positions


def sanitize_label(value: str) -> str:
    return "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in value).strip("._-") or "job"


def make_job_dir(base_output_dir: Path, workbook_path: Path) -> Path:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    label = sanitize_label(f"{workbook_path.stem}_ns5b_gt_resistance_profile")
    job_dir = base_output_dir / f"{label}_{timestamp}"
    job_dir.mkdir(parents=True, exist_ok=True)
    return job_dir


def format_freq(value: float) -> str:
    return format(value, ".2g") if value >= 1.0 else format(value, ".1g")


def load_consensus_by_gt(json_path: Path) -> dict[str, str]:
    rows = json.loads(json_path.read_text(encoding="utf-8"))
    consensus: dict[str, str] = {}
    for row in rows:
        name = str(row.get("name", ""))
        match = re.fullmatch(r"HCV([1-8])NS5B", name)
        if match:
            consensus[match.group(1)] = str(row.get("refSequence", "")).strip().upper()
    return consensus


def load_gt_profile_rows(workbook_path: Path, positions: list[int]) -> dict[str, dict[int, list[tuple[str, float]]]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    result: dict[str, dict[int, list[tuple[str, float]]]] = {}
    wanted = set(positions)
    for sheet_name in wb.sheetnames:
        gt = sheet_name.replace("GT", "")
        ws = wb[sheet_name]
        next(ws.iter_rows(values_only=True))
        by_pos: dict[int, list[tuple[str, float]]] = defaultdict(list)
        for row in ws.iter_rows(min_row=2, values_only=True):
            pos = int(row[0])
            aa = str(row[2])
            pct = float(row[5])
            if pos in wanted:
                by_pos[pos].append((aa, pct))
        result[gt] = by_pos
    wb.close()
    return result


def build_grid(
    consensus_by_gt: dict[str, str],
    profile_rows: dict[str, dict[int, list[tuple[str, float]]]],
    positions: list[int],
) -> list[list[str]]:
    grid: list[list[str]] = []
    for gt in sorted(profile_rows, key=int):
        pos_variants: dict[int, list[str]] = {}
        max_depth = 0
        consensus_seq = consensus_by_gt[gt]
        for pos in positions:
            consensus_aa = consensus_seq[pos - 1]
            variants = [
                f"{aa}-{format_freq(pct)}"
                for aa, pct in sorted(profile_rows[gt].get(pos, []), key=lambda item: (-item[1], item[0]))
                if aa != consensus_aa and aa not in EXCLUDED_AAS and pct > 0.1
            ]
            pos_variants[pos] = variants
            max_depth = max(max_depth, len(variants))

        grid.append([f"GT{gt}"] + [""] * len(positions))
        grid.append(["Position"] + [str(pos) for pos in positions])
        grid.append(["Consensus"] + [consensus_seq[pos - 1] for pos in positions])
        for depth in range(max_depth):
            row = [f"Rank{depth + 1}"]
            for pos in positions:
                variants = pos_variants[pos]
                row.append(variants[depth] if depth < len(variants) else "")
            grid.append(row)
        grid.append([""] + [""] * len(positions))
    return grid


def write_excel(path: Path, grid: list[list[str]], positions: list[int]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "GT_Resistance_Profile"
    gt_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    header_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    consensus_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    bold = Font(bold=True)

    for row_idx, row in enumerate(grid, start=1):
        ws.append(row)
        first = row[0]
        if first.startswith("GT"):
            for cell in ws[row_idx]:
                cell.fill = gt_fill
                cell.font = bold
        elif first == "Position":
            for cell in ws[row_idx]:
                cell.fill = header_fill
                cell.font = bold
        elif first == "Consensus":
            for cell in ws[row_idx]:
                cell.fill = consensus_fill
                cell.font = bold
        for cell in ws[row_idx]:
            cell.alignment = Alignment(horizontal="center")

    for col in range(1, len(positions) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12 if col > 1 else 14
    wb.save(path)


def write_png(path: Path, grid: list[list[str]]) -> None:
    font = ImageFont.load_default()
    row_h = 24
    first_col_w = 110
    other_col_w = 78
    widths = [first_col_w] + [other_col_w] * (len(grid[0]) - 1)
    total_w = sum(widths) + 1
    total_h = len(grid) * row_h + 1
    image = Image.new("RGB", (total_w, total_h), "white")
    draw = ImageDraw.Draw(image)

    y = 0
    for row in grid:
        x = 0
        row_type = row[0]
        for col_idx, text in enumerate(row):
            width = widths[col_idx]
            fill = "white"
            if row_type.startswith("GT"):
                fill = "#D9EAF7"
            elif row_type == "Position":
                fill = "#F2F2F2"
            elif row_type == "Consensus":
                fill = "#E2F0D9"
            draw.rectangle([x, y, x + width, y + row_h], fill=fill, outline="#BFBFBF")
            bbox = draw.textbbox((0, 0), str(text), font=font)
            text_w = bbox[2] - bbox[0]
            text_h = bbox[3] - bbox[1]
            tx = x + (width - text_w) / 2
            ty = y + (row_h - text_h) / 2
            draw.text((tx, ty), str(text), fill="black", font=font)
            x += width
        y += row_h
    image.save(path)


def main() -> int:
    args = parse_args()
    gt_profile_workbook = Path(args.gt_profile_workbook).expanduser()
    gt_aa_json = Path(args.gt_aa_json).expanduser()
    output_dir = Path(args.output_dir)
    positions = parse_positions(args.positions)

    consensus_by_gt = load_consensus_by_gt(gt_aa_json)
    profile_rows = load_gt_profile_rows(gt_profile_workbook, positions)
    grid = build_grid(consensus_by_gt, profile_rows, positions)

    job_dir = make_job_dir(output_dir, gt_profile_workbook)
    excel_path = job_dir / "NS5B_GT_Resistance_Profile_Summary.xlsx"
    png_path = job_dir / "NS5B_GT_Resistance_Profile_Summary.png"
    write_excel(excel_path, grid, positions)
    write_png(png_path, grid)

    summary = {
        "excel": str(excel_path.resolve()),
        "png": str(png_path.resolve()),
        "gene": TARGET_GENE,
        "positions": positions,
        "frequency_threshold_percent": 0.1,
    }
    (job_dir / "summary.json").write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
