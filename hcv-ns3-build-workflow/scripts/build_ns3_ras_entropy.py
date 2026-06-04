#!/usr/bin/env python3

from __future__ import annotations

import argparse
import math
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font


RESISTANCE_POSITIONS = [36, 41, 43, 54, 55, 56, 80, 122, 155, 156, 158, 166, 168, 170, 175]
EXCLUDED_AAS = {"X", "*"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Calculate NS3 RAS-position Shannon entropy by genotype and subtype.")
    parser.add_argument("--gt-profile-workbook", default="outputs/NS3_GT_CompleteProfiles_TabsPerGT.xlsx")
    parser.add_argument("--subtype-profile-workbook", default="outputs/NS3_Subtype_CompleteProfiles_TabsPerGT.xlsx")
    parser.add_argument("--gt-output-xlsx", default="outputs/NS3_GT_RAS_Entropy.xlsx")
    parser.add_argument("--subtype-output-xlsx", default="outputs/NS3_Subtype_RAS_Entropy.xlsx")
    return parser.parse_args()


def shannon_entropy(counts: list[int]) -> float:
    total = sum(counts)
    if total <= 0:
        return 0.0
    entropy = 0.0
    for count in counts:
        if count <= 0:
            continue
        p = count / total
        entropy -= p * math.log2(p)
    return entropy


def round_sig(value: float, digits: int = 2) -> float:
    if value == 0:
        return 0.0
    places = digits - int(math.floor(math.log10(abs(value)))) - 1
    return round(value, places)


def genotype_sort_key(label: str) -> tuple[int, str]:
    if label.startswith("GT") and label[2:].isdigit():
        return int(label[2:]), label
    return 999, label


def subtype_sort_key(label: str) -> tuple[int, str]:
    if label and label[0].isdigit():
        digits = ""
        rest = ""
        for char in label:
            if char.isdigit() and not rest:
                digits += char
            else:
                rest += char
        return int(digits), rest
    return 999, label


def load_gt_counts(workbook_path: Path) -> tuple[dict[str, dict[int, dict[str, int]]], dict[str, dict[int, int]]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    counts: dict[str, dict[int, dict[str, int]]] = defaultdict(lambda: defaultdict(dict))
    coverage: dict[str, dict[int, int]] = defaultdict(dict)
    for sheet_name in wb.sheetnames:
        gt = sheet_name.strip()
        ws = wb[sheet_name]
        header = [str(value) if value is not None else "" for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        index = {name: idx for idx, name in enumerate(header)}
        required = ["NS3Position", "NumSeqsIncludingPosition", "AminoAcid", "CountWithAA"]
        missing = [name for name in required if name not in index]
        if missing:
            raise RuntimeError(f"{workbook_path}:{sheet_name} is missing columns: {', '.join(missing)}")
        for row in ws.iter_rows(min_row=2, values_only=True):
            position = int(row[index["NS3Position"]])
            if position not in RESISTANCE_POSITIONS:
                continue
            aa = str(row[index["AminoAcid"]]).strip().upper()
            if aa in EXCLUDED_AAS:
                continue
            counts[gt][position][aa] = counts[gt][position].get(aa, 0) + int(row[index["CountWithAA"]])
            coverage[gt][position] = int(row[index["NumSeqsIncludingPosition"]])
    wb.close()
    return counts, coverage


def load_subtype_counts(
    workbook_path: Path,
) -> tuple[dict[str, dict[str, dict[int, dict[str, int]]]], dict[str, dict[str, dict[int, int]]]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    counts: dict[str, dict[str, dict[int, dict[str, int]]]] = defaultdict(lambda: defaultdict(lambda: defaultdict(dict)))
    coverage: dict[str, dict[str, dict[int, int]]] = defaultdict(lambda: defaultdict(dict))
    for sheet_name in wb.sheetnames:
        gt = sheet_name.strip()
        ws = wb[sheet_name]
        header = [str(value) if value is not None else "" for value in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        index = {name: idx for idx, name in enumerate(header)}
        required = ["Subtype", "NS3Position", "NumSeqsIncludingPosition", "AminoAcid", "CountWithAA"]
        missing = [name for name in required if name not in index]
        if missing:
            raise RuntimeError(f"{workbook_path}:{sheet_name} is missing columns: {', '.join(missing)}")
        for row in ws.iter_rows(min_row=2, values_only=True):
            subtype = str(row[index["Subtype"]]).strip()
            position = int(row[index["NS3Position"]])
            if position not in RESISTANCE_POSITIONS:
                continue
            aa = str(row[index["AminoAcid"]]).strip().upper()
            if aa in EXCLUDED_AAS:
                continue
            subtype_counts = counts[gt][subtype][position]
            subtype_counts[aa] = subtype_counts.get(aa, 0) + int(row[index["CountWithAA"]])
            coverage[gt][subtype][position] = int(row[index["NumSeqsIncludingPosition"]])
    wb.close()
    return counts, coverage


def add_metadata_sheet(wb: Workbook, source_workbook: Path) -> None:
    ws = wb.create_sheet("metadata")
    rows = [
        ("source_workbook", str(source_workbook.resolve())),
        ("entropy_definition", "Shannon entropy, -sum(p_i * log2(p_i))"),
        ("probability_source", "CountWithAA per amino acid at each RAS position"),
        ("excluded_amino_acids", ",".join(sorted(EXCLUDED_AAS))),
        ("ras_positions", ",".join(str(pos) for pos in RESISTANCE_POSITIONS)),
    ]
    for row in rows:
        ws.append(row)


def style_sheet(ws) -> None:
    ws.freeze_panes = "B2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "General"
            cell.alignment = Alignment(horizontal="center")


def write_gt_entropy(output_path: Path, source_workbook: Path) -> None:
    counts, coverage = load_gt_counts(source_workbook)
    wb = Workbook()
    ws = wb.active
    ws.title = "GT_RAS_Entropy"
    ws.append(["Genotype", *[f"P{pos}" for pos in RESISTANCE_POSITIONS]])
    for gt in sorted(counts, key=genotype_sort_key):
        row = [gt]
        for pos in RESISTANCE_POSITIONS:
            row.append(round_sig(shannon_entropy(list(counts[gt].get(pos, {}).values()))))
        ws.append(row)
    style_sheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_subtype_entropy(output_path: Path, source_workbook: Path) -> None:
    counts, coverage = load_subtype_counts(source_workbook)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for gt in sorted(counts, key=genotype_sort_key):
        ws = wb.create_sheet(gt)
        ws.append(["Subtype", *[f"P{pos}" for pos in RESISTANCE_POSITIONS]])
        for subtype in sorted(counts[gt], key=subtype_sort_key):
            row = [subtype]
            for pos in RESISTANCE_POSITIONS:
                row.append(round_sig(shannon_entropy(list(counts[gt][subtype].get(pos, {}).values()))))
            ws.append(row)
        style_sheet(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    args = parse_args()
    gt_profile_workbook = Path(args.gt_profile_workbook)
    subtype_profile_workbook = Path(args.subtype_profile_workbook)
    gt_output_xlsx = Path(args.gt_output_xlsx)
    subtype_output_xlsx = Path(args.subtype_output_xlsx)

    write_gt_entropy(gt_output_xlsx, gt_profile_workbook)
    write_subtype_entropy(subtype_output_xlsx, subtype_profile_workbook)
    print(f"gt_entropy_xlsx={gt_output_xlsx.resolve()}")
    print(f"subtype_entropy_xlsx={subtype_output_xlsx.resolve()}")
    print(f"ras_positions={','.join(str(pos) for pos in RESISTANCE_POSITIONS)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
