#!/usr/bin/env python3

from __future__ import annotations

import argparse
import re
import subprocess
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook


FASTA_EXTENSIONS = {".fa", ".fasta", ".faa"}
MISSING_AA = {"-", "X", "?", "*"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "For each NS3 genotype, multiple-align subtype amino-acid consensus "
            "sequences, slice aligned positions 36-175, and write one distance "
            "matrix sheet per genotype."
        )
    )
    parser.add_argument("--input-fasta", default="outputs/NS3_Subtype_Consensus.fasta")
    parser.add_argument("--output-xlsx", default="outputs/NS3_Subtype_AA_Distance_Pos36_175.xlsx")
    parser.add_argument("--temp-dir", default="temp/ns3_subtype_aa_distance_matrices")
    parser.add_argument("--start", type=int, default=36, help="1-based aligned start position")
    parser.add_argument("--end", type=int, default=175, help="1-based aligned end position, inclusive")
    parser.add_argument("--mafft-bin", default="mafft")
    return parser.parse_args()


def read_fasta(path: Path) -> list[tuple[str, str]]:
    records: list[tuple[str, str]] = []
    name: str | None = None
    chunks: list[str] = []
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if name is not None:
                    records.append((name, "".join(chunks)))
                name = line[1:].strip().split()[0]
                chunks = []
            else:
                chunks.append(line)
    if name is not None:
        records.append((name, "".join(chunks)))
    if not records:
        raise RuntimeError(f"No FASTA records found in {path}")
    return records


def write_fasta(path: Path, records: list[tuple[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for name, sequence in records:
            handle.write(f">{name}\n")
            for start in range(0, len(sequence), 70):
                handle.write(sequence[start : start + 70] + "\n")


def genotype_sort_key(label: str) -> tuple[int, str]:
    match = re.fullmatch(r"GT(\d+)", label)
    return (int(match.group(1)), label) if match else (999, label)


def subtype_sort_key(label: str) -> tuple[int, str]:
    subtype = label.split("_", 1)[1] if "_" in label else label
    match = re.match(r"(\d+)(.*)", subtype)
    if match:
        return int(match.group(1)), match.group(2)
    return 999, subtype


def group_by_genotype(records: list[tuple[str, str]]) -> dict[str, list[tuple[str, str]]]:
    grouped: dict[str, list[tuple[str, str]]] = defaultdict(list)
    for name, sequence in records:
        if "_" not in name:
            continue
        genotype = name.split("_", 1)[0]
        grouped[genotype].append((name, sequence))
    return {
        genotype: sorted(items, key=lambda item: subtype_sort_key(item[0]))
        for genotype, items in grouped.items()
    }


def run_mafft(input_fasta: Path, aligned_fasta: Path, mafft_bin: str) -> None:
    result = subprocess.run(
        [mafft_bin, "--auto", str(input_fasta)],
        check=True,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    aligned_fasta.write_text(result.stdout, encoding="utf-8")


def pairwise_distance(seq_a: str, seq_b: str) -> tuple[float | None, int, int]:
    differences = 0
    compared = 0
    for aa_a, aa_b in zip(seq_a, seq_b):
        if aa_a.upper() in MISSING_AA or aa_b.upper() in MISSING_AA:
            continue
        compared += 1
        if aa_a != aa_b:
            differences += 1
    denominator = len(seq_a)
    if denominator == 0:
        return None, differences, compared
    return differences / denominator, differences, compared


def display_name(label: str) -> str:
    return label.split("_", 1)[1] if "_" in label else label


def add_distance_sheet(wb: Workbook, genotype: str, records: list[tuple[str, str]]) -> None:
    labels = [name for name, _ in records]
    display_labels = [display_name(name) for name in labels]
    seqs = dict(records)
    ws = wb.create_sheet(genotype)
    ws.append(["Subtype", *display_labels[1:]])
    for row_index, row_name in enumerate(labels[:-1]):
        row: list[str | float | None] = [display_name(row_name)]
        for col_name in labels[1:]:
            col_index = labels.index(col_name)
            if row_index < col_index:
                distance, _, _ = pairwise_distance(seqs[row_name], seqs[col_name])
                row.append(distance)
            else:
                row.append(None)
        ws.append(row)
    ws.freeze_panes = "B2"
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = "0.0%"


def write_summary(path: Path, rows: list[dict[str, str | int]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    headers = ["genotype", "subtype", "input_length", "status", "reason"]
    with path.open("w", encoding="utf-8") as handle:
        handle.write(",".join(headers) + "\n")
        for row in rows:
            handle.write(",".join(str(row.get(header, "")) for header in headers) + "\n")


def main() -> int:
    args = parse_args()
    input_fasta = Path(args.input_fasta)
    output_xlsx = Path(args.output_xlsx)
    temp_dir = Path(args.temp_dir)
    if args.start < 1 or args.end < args.start:
        raise SystemExit("--start/--end must define a valid 1-based inclusive range")
    if input_fasta.suffix.lower() not in FASTA_EXTENSIONS:
        raise SystemExit(f"Input FASTA extension should be one of {sorted(FASTA_EXTENSIONS)}")

    records = read_fasta(input_fasta)
    grouped = group_by_genotype(records)
    temp_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    summary_rows: list[dict[str, str | int]] = []
    sheet_count = 0

    for genotype in sorted(grouped, key=genotype_sort_key):
        group_records = grouped[genotype]
        kept = []
        for name, sequence in group_records:
            subtype = display_name(name)
            if len(sequence) < args.end:
                summary_rows.append(
                    {
                        "genotype": genotype,
                        "subtype": subtype,
                        "input_length": len(sequence),
                        "status": "ignored",
                        "reason": f"input length below aligned end {args.end}",
                    }
                )
                continue
            kept.append((name, sequence))
            summary_rows.append(
                {
                    "genotype": genotype,
                    "subtype": subtype,
                    "input_length": len(sequence),
                    "status": "included",
                    "reason": "",
                }
            )
        if len(kept) < 2:
            continue

        group_input = temp_dir / f"{genotype}_subtype_consensus.fasta"
        group_aligned = temp_dir / f"{genotype}_subtype_consensus_aligned.fasta"
        write_fasta(group_input, kept)
        run_mafft(group_input, group_aligned, args.mafft_bin)
        aligned = read_fasta(group_aligned)
        lengths = {len(sequence) for _, sequence in aligned}
        if len(lengths) != 1:
            raise RuntimeError(f"{genotype} MAFFT output has inconsistent lengths: {sorted(lengths)}")
        aligned_length = next(iter(lengths))
        if aligned_length < args.end:
            continue
        window_records = [
            (name, sequence[args.start - 1 : args.end])
            for name, sequence in aligned
        ]
        add_distance_sheet(wb, genotype, window_records)
        sheet_count += 1

    if sheet_count == 0:
        raise RuntimeError("No genotype groups had at least two subtype records covering the requested range")

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    summary_path = temp_dir / f"{output_xlsx.stem}_summary.csv"
    write_summary(summary_path, summary_rows)

    print(f"output_xlsx={output_xlsx.resolve()}")
    print(f"temp_dir={temp_dir.resolve()}")
    print(f"summary_csv={summary_path.resolve()}")
    print(f"sheet_count={sheet_count}")
    print(f"input_record_count={len(records)}")
    print(f"ignored_record_count={sum(1 for row in summary_rows if row['status'] == 'ignored')}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
