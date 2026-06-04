#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import subprocess
import tempfile
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


BLAST_OUTFMT = "6 qseqid sseqid length mismatch gaps pident evalue bitscore qstart qend sstart send qseq sseq"
FASTA_EXTENSIONS = {".fa", ".fasta", ".fna"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build a combined NS3 subtype assignment workbook by aligning study sequences "
            "to genotype-matched subtype genome references."
        )
    )
    parser.add_argument("--combined-workbook", required=True, help="Path to NS3_Alignments_combined.xlsx")
    parser.add_argument("--fasta-dir", required=True, help="Directory containing study FASTA files")
    parser.add_argument("--subtype-json", required=True, help="Path to HCV_Subtype_Refs_By_Genome_NA.json")
    parser.add_argument("--output-dir", default="outputs", help="Base output directory")
    parser.add_argument("--min-aligned-nt", type=int, default=200, help="Skip hits shorter than this overlap length")
    parser.add_argument(
        "--genotype-subtype-csv",
        help=(
            "Optional included_accessions_genotype_subtype.csv from the metadata-filter step. "
            "Rows with subtype use that value as ClosestSubtype instead of distance selection."
        ),
    )
    return parser.parse_args()


def sanitize_label(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return text.strip("._-") or "job"


def script_temp_dir() -> Path:
    path = Path("temp") / "hcv-ns3-build-workflow" / Path(__file__).stem
    path.mkdir(parents=True, exist_ok=True)
    return path


def make_job_dir(base_output_dir: Path, workbook_path: Path) -> Path:
    label = sanitize_label(f"{workbook_path.stem}_ns3_subtype_distance")
    return Path(tempfile.mkdtemp(prefix=f"{label}_", dir=script_temp_dir()))


def parse_fasta(path: Path) -> list[tuple[str, str]]:
    records: list[tuple[str, str]] = []
    header: str | None = None
    chunks: list[str] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line:
            continue
        if line.startswith(">"):
            if header is not None:
                records.append((header, "".join(chunks).upper()))
            header = line[1:].strip()
            chunks = []
        else:
            chunks.append(re.sub(r"\s+", "", line))
    if header is not None:
        records.append((header, "".join(chunks).upper()))
    return records


def accession_from_header(header: str) -> str:
    return header.split()[0]


def normalize_genotype(value: str) -> str:
    match = re.match(r"([1-8])", value.strip())
    return match.group(1) if match else ""


def normalize_subtype(value: str, genotype: str = "") -> str:
    text = value.strip().lower()
    if re.fullmatch(r"[a-z][a-z0-9]*", text) and genotype:
        return f"{genotype}{text}"
    return text


def default_genotype_subtype_csv(fasta_dir: Path) -> Path:
    return fasta_dir.parent / "included_accessions_genotype_subtype.csv"


def load_metadata_assignments(path: Path | None) -> dict[str, dict[str, str]]:
    if path is None or not path.is_file():
        return {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"accession", "genotype", "subtype", "column_name"}
        missing = required.difference(reader.fieldnames or [])
        if missing:
            raise RuntimeError(f"Columns missing from {path}: {', '.join(sorted(missing))}")
        assignments: dict[str, dict[str, str]] = {}
        for row in reader:
            accession = (row.get("accession") or "").strip()
            if not accession:
                continue
            genotype = normalize_genotype(row.get("genotype") or "")
            assignments[accession] = {
                "genotype": genotype,
                "subtype": normalize_subtype(row.get("subtype") or "", genotype),
                "column_name": (row.get("column_name") or "").strip(),
            }
    return assignments


CODON_TABLE = {
    "TTT": "F",
    "TTC": "F",
    "TTA": "L",
    "TTG": "L",
    "TCT": "S",
    "TCC": "S",
    "TCA": "S",
    "TCG": "S",
    "TAT": "Y",
    "TAC": "Y",
    "TAA": "*",
    "TAG": "*",
    "TGT": "C",
    "TGC": "C",
    "TGA": "*",
    "TGG": "W",
    "CTT": "L",
    "CTC": "L",
    "CTA": "L",
    "CTG": "L",
    "CCT": "P",
    "CCC": "P",
    "CCA": "P",
    "CCG": "P",
    "CAT": "H",
    "CAC": "H",
    "CAA": "Q",
    "CAG": "Q",
    "CGT": "R",
    "CGC": "R",
    "CGA": "R",
    "CGG": "R",
    "ATT": "I",
    "ATC": "I",
    "ATA": "I",
    "ATG": "M",
    "ACT": "T",
    "ACC": "T",
    "ACA": "T",
    "ACG": "T",
    "AAT": "N",
    "AAC": "N",
    "AAA": "K",
    "AAG": "K",
    "AGT": "S",
    "AGC": "S",
    "AGA": "R",
    "AGG": "R",
    "GTT": "V",
    "GTC": "V",
    "GTA": "V",
    "GTG": "V",
    "GCT": "A",
    "GCC": "A",
    "GCA": "A",
    "GCG": "A",
    "GAT": "D",
    "GAC": "D",
    "GAA": "E",
    "GAG": "E",
    "GGT": "G",
    "GGC": "G",
    "GGA": "G",
    "GGG": "G",
}


def filename_matches_refid(filename: str, refid: str) -> bool:
    return filename.startswith(f"{refid}_")


def load_combined_rows(workbook_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [cell for cell in next(ws.iter_rows(values_only=True))]
    index = {str(name): idx for idx, name in enumerate(header)}
    required = ["RefID", "RefName", "GenBankAccession", "BestGT"]
    for name in required:
        if name not in index:
            raise RuntimeError(f"Column '{name}' not found in {workbook_path}")

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        refid = str(row[index["RefID"]]).strip()
        if not refid:
            continue
        best_gt = str(row[index["BestGT"]]).strip()
        accession = str(row[index["GenBankAccession"]]).strip()
        if not best_gt or not accession:
            continue
        rows.append(
            {
                "RefID": refid,
                "RefName": str(row[index["RefName"]]).strip(),
                "AccessionID": accession,
                "ClosestGT": best_gt,
            }
        )
    wb.close()
    return rows


def load_subtype_references(json_path: Path) -> dict[str, list[dict[str, str]]]:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    refs_by_gt: dict[str, list[dict[str, str]]] = defaultdict(list)
    for rec in data:
        genotype_name = str(rec.get("genotypeName", "")).strip()
        match = re.search(r"Genotype\s*([0-9]+[A-Za-z0-9]*)", genotype_name) or re.search(
            r"Genotype([0-9]+[A-Za-z0-9]*)", genotype_name
        )
        if not match:
            continue
        subtype = match.group(1)
        gt_match = re.match(r"(\d+)", subtype)
        if not gt_match:
            continue
        gt = gt_match.group(1)
        accession = str(rec.get("accession", "")).strip() or subtype
        sequence = str(rec.get("sequence", "")).strip().upper()
        if not sequence:
            continue
        refs_by_gt[gt].append(
            {
                "subtype": subtype,
                "accession": accession,
                "sequence": sequence,
            }
        )
    return refs_by_gt


def build_refid_to_fasta(fasta_dir: Path) -> dict[str, Path]:
    mapping: dict[str, Path] = {}
    for path in sorted(fasta_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in FASTA_EXTENSIONS:
            continue
        name = path.name
        if "_" not in name:
            continue
        refid = name.split("_", 1)[0]
        if refid in mapping:
            raise RuntimeError(f"Found multiple FASTA files for RefID {refid}")
        mapping[refid] = path
    return mapping


def write_fasta_entries(path: Path, entries: list[tuple[str, str]]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for header, sequence in entries:
            handle.write(f">{header}\n")
            for start in range(0, len(sequence), 70):
                handle.write(sequence[start : start + 70] + "\n")


def build_subtype_db(job_dir: Path, gt: str, refs: list[dict[str, str]]) -> tuple[Path, dict[str, dict[str, str]]]:
    fasta_path = job_dir / f"ns3_subtype_gt{gt}.fasta"
    entries: list[tuple[str, str]] = []
    subject_meta: dict[str, dict[str, str]] = {}
    for idx, ref in enumerate(refs, start=1):
        subject_id = f"GT{gt}_REF{idx}"
        subject_meta[subject_id] = {
            "subtype": ref["subtype"],
            "accession": ref["accession"],
        }
        entries.append((subject_id, ref["sequence"]))
    write_fasta_entries(fasta_path, entries)
    db_prefix = job_dir / f"ns3_subtype_gt{gt}_db"
    subprocess.run(
        [
            "makeblastdb",
            "-in",
            str(fasta_path),
            "-dbtype",
            "nucl",
            "-out",
            str(db_prefix),
            "-parse_seqids",
        ],
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        text=True,
    )
    return db_prefix, subject_meta


def run_blastn(query_path: Path, db_prefix: Path, out_path: Path) -> list[dict[str, Any]]:
    subprocess.run(
        [
            "blastn",
            "-query",
            str(query_path),
            "-db",
            str(db_prefix),
            "-dust",
            "no",
            "-task",
            "blastn",
            "-evalue",
            "1e-6",
            "-max_hsps",
            "1",
            "-max_target_seqs",
            "100",
            "-outfmt",
            BLAST_OUTFMT,
            "-out",
            str(out_path),
        ],
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        text=True,
    )
    hits: list[dict[str, Any]] = []
    for line in out_path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        p = line.split("\t")
        length = int(p[2])
        mismatch = int(p[3])
        gaps = int(p[4])
        hits.append(
            {
                "qseqid": p[0],
                "sseqid": p[1],
                "length": length,
                "mismatch": mismatch,
                "gaps": gaps,
                "bitscore": float(p[7]),
                "qstart": int(p[8]),
                "qend": int(p[9]),
                "sstart": int(p[10]),
                "send": int(p[11]),
                "aligned_query_nt": p[12],
                "aligned_subject_nt": p[13],
                "distance": (mismatch + gaps) / length if length else None,
            }
        )
    try:
        out_path.unlink()
    except OSError:
        pass
    return hits


def choose_best_by_subtype(
    hits: list[dict[str, Any]],
    subject_meta: dict[str, dict[str, str]],
    min_aligned_nt: int,
) -> dict[str, list[dict[str, Any]]]:
    best: dict[tuple[str, str], dict[str, Any]] = {}
    for hit in hits:
        meta = subject_meta.get(hit["sseqid"])
        if meta is None:
            continue
        subtype = meta["subtype"]
        key = (hit["qseqid"], subtype)
        current = best.get(key)
        candidate = {
            "subtype": subtype,
            "subtype_ref_accession": meta["accession"],
            "distance": hit["distance"],
            "aligned_nt": hit["length"],
            "bitscore": hit["bitscore"],
            "qstart": hit["qstart"],
            "qend": hit["qend"],
            "sstart": hit["sstart"],
            "send": hit["send"],
            "aligned_query_nt": hit["aligned_query_nt"],
            "aligned_subject_nt": hit["aligned_subject_nt"],
        }
        if current is None or (
            candidate["distance"],
            -candidate["aligned_nt"],
            -candidate["bitscore"],
        ) < (
            current["distance"],
            -current["aligned_nt"],
            -current["bitscore"],
        ):
            best[key] = candidate

    by_query: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for (qseqid, _subtype), hit in best.items():
        by_query[qseqid].append(hit)
    for qseqid in by_query:
        by_query[qseqid].sort(key=lambda item: (item["distance"], -item["aligned_nt"], -item["bitscore"]))
    return by_query


def normalize_nt(nt: str) -> str:
    return re.sub(r"[^ACGTRYSWKMBDHVN-]", "N", nt.upper())


def build_aa_marker(reference: str, query: str) -> str:
    chars: list[str] = []
    for ref_char, query_char in zip(reference, query):
        if ref_char == "-" or query_char == "-":
            chars.append(" ")
        elif ref_char == query_char:
            chars.append("|")
        else:
            chars.append(".")
    return "".join(chars)


def translate_nt(sequence: str) -> str:
    aa: list[str] = []
    for start in range(0, len(sequence), 3):
        codon = sequence[start : start + 3]
        if len(codon) < 3:
            break
        if any(base not in "ACGT" for base in codon):
            aa.append("X")
        else:
            aa.append(CODON_TABLE.get(codon, "X"))
    return "".join(aa)


def translate_aligned_nt(sequence: str) -> str:
    aa: list[str] = []
    for start in range(0, len(sequence), 3):
        codon = sequence[start : start + 3]
        if len(codon) < 3:
            break
        if codon == "---":
            aa.append("-")
        elif "-" in codon or any(base not in "ACGT" for base in codon):
            aa.append("X")
        else:
            aa.append(CODON_TABLE.get(codon, "X"))
    return "".join(aa)


def extract_aa_window(sequence: str, hit: dict[str, Any]) -> tuple[int, int, str]:
    qstart = min(int(hit["qstart"]), int(hit["qend"]))
    qend = max(int(hit["qstart"]), int(hit["qend"]))
    sstart = min(int(hit["sstart"]), int(hit["send"]))
    send = max(int(hit["sstart"]), int(hit["send"]))

    start_aa = ((sstart - 1) // 3) + 1
    end_aa = send // 3

    leading_trim = (3 - ((sstart - 1) % 3)) % 3
    usable_start = qstart + leading_trim
    usable_end = qend - ((qend - usable_start + 1) % 3)
    if usable_end < usable_start:
        return start_aa, start_aa - 1, ""

    nt_window = normalize_nt(sequence[usable_start - 1 : usable_end])
    aa_sequence = translate_nt(nt_window)
    end_aa = start_aa + len(aa_sequence) - 1 if aa_sequence else start_aa - 1
    return start_aa, end_aa, aa_sequence


def write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "NS3_Subtype_Distance"
    fieldnames = [
        "RefID",
        "RefName",
        "AccessionID",
        "ClosestGT",
        "ClosestSubtype",
        "ClosestSubtypeAssignmentSource",
        "ClosestSubtypeMetadataColumn",
        "ClosestSubtypeRefAccession",
        "ClosestSubtypeDistance",
        "NextClosestSubtype",
        "NextClosestSubtypeDistance",
        "AlignedNT",
        "NextClosestSubtypeAlignedNT",
        "StartAAPosition",
        "EndAAPosition",
        "AASequence",
    ]
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])
    workbook.save(path)


def write_unassigned_report(path: Path, reason_rows: list[tuple[str, dict[str, Any]]]) -> None:
    fieldnames = ["AccessionID", "Reason", "RefID", "RefName", "ClosestGT"]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for reason, row in reason_rows:
            writer.writerow(
                {
                    "Reason": reason,
                    "RefID": row.get("RefID", ""),
                    "RefName": row.get("RefName", ""),
                    "AccessionID": row.get("AccessionID", ""),
                    "ClosestGT": row.get("ClosestGT", ""),
                }
            )


def write_missing_accession_reason_files(
    temp_dir: Path,
    reason_rows: list[tuple[str, dict[str, Any]]],
) -> Path:
    fieldnames = ["Reason", "RefID", "RefName", "AccessionID", "ClosestGT"]
    all_reasons_path = temp_dir / "missing_accession_reason.csv"

    with all_reasons_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(fieldnames)
        for reason, row in reason_rows:
            writer.writerow([row.get("AccessionID", ""), reason, row.get("RefID", ""), row.get("RefName", ""), row.get("ClosestGT", "")])
    return all_reasons_path


def cleanup_db_files(job_dir: Path) -> None:
    for path in job_dir.iterdir():
        if path.name.startswith("ns3_subtype_gt") and path.suffix in {
            ".fasta",
            ".nhr",
            ".nin",
            ".nsq",
            ".ndb",
            ".not",
            ".ntf",
            ".nto",
        }:
            try:
                path.unlink()
            except OSError:
                pass


def main() -> int:
    args = parse_args()
    combined_workbook = Path(args.combined_workbook).expanduser()
    fasta_dir = Path(args.fasta_dir).expanduser()
    subtype_json = Path(args.subtype_json).expanduser()
    output_dir = Path(args.output_dir)
    metadata_csv = (
        Path(args.genotype_subtype_csv).expanduser()
        if args.genotype_subtype_csv
        else default_genotype_subtype_csv(fasta_dir)
    )
    metadata_assignments = load_metadata_assignments(metadata_csv)

    base_rows = load_combined_rows(combined_workbook)
    refs_by_gt = load_subtype_references(subtype_json)
    refid_to_fasta = build_refid_to_fasta(fasta_dir)
    job_dir = make_job_dir(output_dir, combined_workbook)
    output_path = output_dir / "NS3_Subtype_AllStudies_WSeqs.xlsx"

    rows_by_refid: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in base_rows:
        rows_by_refid[row["RefID"]].append(row)

    query_entries_by_gt: dict[str, list[tuple[str, str]]] = defaultdict(list)
    entries_by_qseqid: dict[str, str] = {}
    row_lookup: dict[str, dict[str, Any]] = {}
    skipped_missing_fasta: list[dict[str, str]] = []
    skipped_missing_sequence: list[dict[str, str]] = []
    skipped_missing_subtype_refs: list[dict[str, str]] = []
    skipped_no_subtype_hit: list[dict[str, str]] = []
    output_rows: list[dict[str, Any]] = []

    for refid, rows in rows_by_refid.items():
        fasta_path = refid_to_fasta.get(refid)
        if fasta_path is None:
            for row in rows:
                skipped_missing_fasta.append(row)
            continue
        sequence_by_accession = {accession_from_header(h): seq for h, seq in parse_fasta(fasta_path)}
        for row in rows:
            accession = row["AccessionID"]
            sequence = sequence_by_accession.get(accession)
            if not sequence:
                skipped_missing_sequence.append(row)
                continue
            gt = row["ClosestGT"]
            metadata = metadata_assignments.get(accession, {})
            metadata_subtype = normalize_subtype(metadata.get("subtype", ""), gt)
            if metadata_subtype:
                output_rows.append(
                    {
                        "RefID": row["RefID"],
                        "RefName": row["RefName"],
                        "AccessionID": row["AccessionID"],
                        "ClosestGT": gt,
                        "ClosestSubtype": metadata_subtype,
                        "ClosestSubtypeAssignmentSource": "metadata",
                        "ClosestSubtypeMetadataColumn": metadata.get("column_name", ""),
                        "ClosestSubtypeRefAccession": "",
                        "ClosestSubtypeDistance": "",
                        "NextClosestSubtype": "",
                        "NextClosestSubtypeDistance": "",
                        "AlignedNT": "",
                        "NextClosestSubtypeAlignedNT": "",
                        "StartAAPosition": "",
                        "EndAAPosition": "",
                        "AASequence": "",
                        "AlignedQueryNT": "",
                        "AlignedSubjectNT": "",
                    }
                )
                continue
            qseqid = f"{refid}|{accession}"
            query_entries_by_gt[gt].append((qseqid, sequence))
            entries_by_qseqid[qseqid] = sequence
            row_lookup[qseqid] = row

    for gt, entries in sorted(query_entries_by_gt.items(), key=lambda item: int(item[0])):
        refs = refs_by_gt.get(gt, [])
        if not refs:
            for qseqid, _sequence in entries:
                skipped_missing_subtype_refs.append(row_lookup[qseqid])
            continue
        query_fasta = job_dir / f"ns3_queries_gt{gt}.fasta"
        write_fasta_entries(query_fasta, entries)
        db_prefix, subject_meta = build_subtype_db(job_dir, gt, refs)
        hits = run_blastn(query_fasta, db_prefix, job_dir / f"ns3_gt{gt}.blast.tsv")
        hits_by_query = choose_best_by_subtype(hits, subject_meta, args.min_aligned_nt)

        for qseqid, row in row_lookup.items():
            if row["ClosestGT"] != gt:
                continue
            subtype_hits = hits_by_query.get(qseqid)
            if not subtype_hits:
                skipped_no_subtype_hit.append(row)
                continue
            best = subtype_hits[0]
            second = subtype_hits[1] if len(subtype_hits) > 1 else None
            start_aa, end_aa, aa_sequence = extract_aa_window(entries_by_qseqid[qseqid], best)
            output_rows.append(
                {
                    "RefID": row["RefID"],
                    "RefName": row["RefName"],
                    "AccessionID": row["AccessionID"],
                    "ClosestGT": row["ClosestGT"],
                    "ClosestSubtype": best["subtype"],
                    "ClosestSubtypeAssignmentSource": "distance",
                    "ClosestSubtypeMetadataColumn": "",
                    "ClosestSubtypeRefAccession": best["subtype_ref_accession"],
                    "ClosestSubtypeDistance": best["distance"],
                    "NextClosestSubtype": second["subtype"] if second else "",
                    "NextClosestSubtypeDistance": second["distance"] if second else "",
                    "AlignedNT": best["aligned_nt"],
                    "NextClosestSubtypeAlignedNT": second["aligned_nt"] if second else "",
                    "StartAAPosition": start_aa if aa_sequence else "",
                    "EndAAPosition": end_aa if aa_sequence else "",
                    "AASequence": aa_sequence,
                    "AlignedQueryNT": best["aligned_query_nt"],
                    "AlignedSubjectNT": best["aligned_subject_nt"],
                }
            )
        try:
            query_fasta.unlink()
        except OSError:
            pass

    temp_output_dir = script_temp_dir()
    unassigned_report = temp_output_dir / "unassigned_subtype_accessions.csv"
    unassigned_reason_rows = (
        [("missing_fasta", row) for row in skipped_missing_fasta]
        + [("missing_sequence", row) for row in skipped_missing_sequence]
        + [("missing_subtype_refs", row) for row in skipped_missing_subtype_refs]
        + [("no_subtype_hit", row) for row in skipped_no_subtype_hit]
    )
    write_unassigned_report(unassigned_report, unassigned_reason_rows)
    missing_accession_reason_file = write_missing_accession_reason_files(
        temp_output_dir,
        unassigned_reason_rows,
    )

    output_rows.sort(key=lambda row: (int(row["RefID"]), row["AccessionID"]))
    write_xlsx(output_path, output_rows)
    # Historical extra output kept for reference only.
    # (job_dir / "workflow_request.txt").write_text(
    #     Path("notes/ns3_subtype_distance_workflow_2026-05-11.md").read_text(encoding="utf-8"),
    #     encoding="utf-8",
    # )
    summary = {
        "output_workbook": str(output_path.resolve()),
        "row_count": len(output_rows),
        "skipped_missing_fasta": len(skipped_missing_fasta),
        "skipped_missing_sequence": len(skipped_missing_sequence),
        "skipped_missing_subtype_refs": len(skipped_missing_subtype_refs),
        "skipped_no_subtype_hit": len(skipped_no_subtype_hit),
        "unassigned_subtype_report": str(unassigned_report.resolve()),
        "missing_accession_reason_file": str(missing_accession_reason_file.resolve()),
        "input_row_count": len(base_rows),
        "metadata_assignment_csv": str(metadata_csv.resolve()) if metadata_csv.is_file() else "",
        "metadata_assignment_count": len(metadata_assignments),
        "metadata_closest_subtype_count": sum(
            1 for row in output_rows if row.get("ClosestSubtypeAssignmentSource") == "metadata"
        ),
    }
    # Historical extra output kept for reference only.
    # (job_dir / "summary.json").write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
    cleanup_db_files(job_dir)
    shutil.rmtree(job_dir, ignore_errors=True)
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
