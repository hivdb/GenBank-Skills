#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_FASTA_EXTENSIONS = {
    ".fa",
    ".faa",
    ".fasta",
    ".fna",
    ".fas",
    ".ffn",
    ".frn",
    ".seq",
}
KNOWN_QUASISPECIES_REFIDS = {
    "19",
    "31",
    "32",
    "34",
    "70",
    "81",
    "115",
    "262",
    "1044",
    "2043",
    "2071",
    "2129",
    "2139",
    "2175",
    "2195",
    "2212",
    "2216",
    "2225",
    "2324",
}

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Read one Excel worksheet, keep rows where NumPatients > 0, "
            "collect RefIDs, and find matching FASTA files by filename prefix."
        )
    )
    parser.add_argument("--excel-file", required=True, help="Path to the Excel workbook")
    parser.add_argument("--sheet", required=True, help="Worksheet name to read")
    parser.add_argument("--fasta-dir", required=True, help="Directory containing FASTA files")
    parser.add_argument("--output-dir", default="outputs", help="Base output directory")
    parser.add_argument("--refid-column", default="RefID", help="Column name holding RefID values")
    parser.add_argument(
        "--numpatients-column",
        default="NumPatients",
        help="Column name holding the NumPatients values",
    )
    parser.add_argument(
        "--positive-column",
        action="append",
        default=[],
        help="Additional column that must be numeric and greater than 0; repeat for multiple columns",
    )
    parser.add_argument(
        "--exclude-refid",
        action="append",
        default=[],
        help="RefID to exclude after filtering; repeat for multiple RefIDs",
    )
    parser.add_argument(
        "--exclude-known-quasispecies-refids",
        action="store_true",
        default=True,
        help="Exclude the built-in RefID set used by downstream genotype/subtype workflows (default: enabled)",
    )
    return parser.parse_args()


def sanitize_label(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return text.strip("._-") or "job"


def make_job_dir(base_output_dir: Path, excel_file: Path, sheet_name: str) -> Path:
    label = sanitize_label(f"{excel_file.stem}_{sheet_name}")
    job_dir = base_output_dir / f"refid_fasta_{label}"
    if job_dir.exists():
        shutil.rmtree(job_dir)
    job_dir.mkdir(parents=True, exist_ok=True)
    return job_dir


def normalize_header(value: Any) -> str:
    return str(value).strip()


def parse_positive_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def load_matching_refids(
    excel_file: Path,
    sheet_name: str,
    refid_column: str,
    numpatients_column: str,
    positive_columns: list[str],
    excluded_refids: set[str],
) -> tuple[list[str], list[dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(excel_file, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise RuntimeError(f"Worksheet '{sheet_name}' was not found in {excel_file}")

    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError(f"Worksheet '{sheet_name}' is empty")

    headers = [normalize_header(value) for value in rows[0]]
    header_index = {header: idx for idx, header in enumerate(headers) if header}

    if refid_column not in header_index:
        raise RuntimeError(f"Column '{refid_column}' was not found in worksheet '{sheet_name}'")
    if numpatients_column not in header_index:
        raise RuntimeError(f"Column '{numpatients_column}' was not found in worksheet '{sheet_name}'")
    for column in positive_columns:
        if column not in header_index:
            raise RuntimeError(f"Column '{column}' was not found in worksheet '{sheet_name}'")

    refid_idx = header_index[refid_column]
    numpatients_idx = header_index[numpatients_column]
    positive_indices = {column: header_index[column] for column in positive_columns}

    matching_refids: list[str] = []
    matching_rows: list[dict[str, Any]] = []
    scanned_rows = 0
    qualifying_rows = 0
    skipped_non_numeric = 0
    skipped_missing_refid = 0
    skipped_additional_positive_filter = 0
    skipped_excluded_refid = 0

    for row in rows[1:]:
        scanned_rows += 1
        number = parse_positive_number(row[numpatients_idx] if numpatients_idx < len(row) else None)
        if number is None:
            skipped_non_numeric += 1
            continue
        if number <= 0:
            continue

        failed_positive_filter = False
        positive_values: dict[str, float] = {}
        for column, idx in positive_indices.items():
            extra_value = parse_positive_number(row[idx] if idx < len(row) else None)
            if extra_value is None or extra_value <= 0:
                failed_positive_filter = True
                break
            positive_values[column] = extra_value
        if failed_positive_filter:
            skipped_additional_positive_filter += 1
            continue

        refid_raw = row[refid_idx] if refid_idx < len(row) else None
        refid = str(refid_raw).strip() if refid_raw is not None else ""
        if not refid:
            skipped_missing_refid += 1
            continue
        if refid in excluded_refids:
            skipped_excluded_refid += 1
            continue

        qualifying_rows += 1
        matching_refids.append(refid)
        matching_rows.append(
            {
                "refid": refid,
                numpatients_column: number,
                **positive_values,
            }
        )

    diagnostics = {
        "worksheet": sheet_name,
        "headers": headers,
        "rows_scanned": scanned_rows,
        "qualifying_rows": qualifying_rows,
        "skipped_non_numeric_numpatients": skipped_non_numeric,
        "skipped_missing_refid": skipped_missing_refid,
        "skipped_additional_positive_filter": skipped_additional_positive_filter,
        "skipped_excluded_refid": skipped_excluded_refid,
    }
    return matching_refids, matching_rows, diagnostics


def looks_like_fasta(path: Path) -> bool:
    return path.is_file() and path.suffix.lower() in DEFAULT_FASTA_EXTENSIONS


def filename_matches_refid(filename: str, refid: str) -> bool:
    if filename == refid:
        return True
    prefix = f"{refid}_"
    return filename.startswith(prefix)


def find_matching_files(fasta_dir: Path, refids: list[str]) -> tuple[list[str], dict[str, list[str]]]:
    files_by_refid: dict[str, list[str]] = {refid: [] for refid in refids}
    for path in sorted(fasta_dir.rglob("*")):
        if not looks_like_fasta(path):
            continue
        filename = path.name
        for refid in refids:
            if filename_matches_refid(filename, refid):
                files_by_refid[refid].append(str(path))

    matched_files: list[str] = []
    for refid in refids:
        matched_files.extend(files_by_refid[refid])
    return matched_files, files_by_refid


def write_lines(path: Path, lines: list[str]) -> None:
    path.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")


def main() -> int:
    args = parse_args()
    excel_file = Path(args.excel_file).expanduser()
    fasta_dir = Path(args.fasta_dir).expanduser()
    base_output_dir = Path(args.output_dir)

    if not excel_file.exists():
        raise RuntimeError(f"Excel file was not found: {excel_file}")
    if not fasta_dir.is_dir():
        raise RuntimeError(f"FASTA directory was not found or is not a directory: {fasta_dir}")

    excluded_refids = set(args.exclude_refid)
    if args.exclude_known_quasispecies_refids:
        excluded_refids.update(KNOWN_QUASISPECIES_REFIDS)

    base_output_dir.mkdir(parents=True, exist_ok=True)
    job_dir = make_job_dir(base_output_dir, excel_file, args.sheet)

    refids, matching_rows, diagnostics = load_matching_refids(
        excel_file,
        args.sheet,
        args.refid_column,
        args.numpatients_column,
        args.positive_column,
        excluded_refids,
    )
    matched_files, files_by_refid = find_matching_files(fasta_dir, refids)

    unmatched_refids = [refid for refid in refids if not files_by_refid[refid]]
    matched_filenames = [Path(path).name for path in matched_files]

    write_lines(job_dir / "matching_refids.txt", refids)
    write_lines(job_dir / "matched_fasta_files.txt", matched_files)

    summary = {
        "excel_file": str(excel_file.resolve()),
        "sheet": args.sheet,
        "fasta_dir": str(fasta_dir.resolve()),
        "refid_column": args.refid_column,
        "numpatients_column": args.numpatients_column,
        "excluded_refids": sorted(excluded_refids),
        "refid_count": len(refids),
        "matched_file_count": len(matched_files),
        "matched_filenames": matched_filenames,
        "refids": refids,
        "matching_rows": matching_rows,
        "unmatched_refids": unmatched_refids,
        "files_by_refid": files_by_refid,
        "diagnostics": diagnostics,
        "output_dir": str(job_dir.resolve()),
    }
    write_json(job_dir / "summary.json", summary)

    print(json.dumps(summary, indent=2, ensure_ascii=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
