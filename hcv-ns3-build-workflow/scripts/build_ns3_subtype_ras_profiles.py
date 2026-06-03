#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill


RESISTANCE_POSITIONS = [36, 41, 43, 54, 55, 56, 80, 122, 155, 156, 158, 166, 168, 170, 175]
EXCLUDED_AAS = {"X", "*"}
FREQUENCY_THRESHOLD_PERCENT = 1.0
type VariantCell = list[tuple[str, str]]
type GridCell = str | int | VariantCell


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build subtype-level NS3 resistance-position AA profile summary in Excel."
    )
    parser.add_argument("--subtype-profile-workbook", required=True)
    parser.add_argument("--gt-aa-json", required=True)
    parser.add_argument("--output-dir", default="outputs")
    return parser.parse_args()


def script_temp_dir() -> Path:
    path = Path("temp") / "hcv-ns3-build-workflow" / Path(__file__).stem
    path.mkdir(parents=True, exist_ok=True)
    return path


def sanitize_label(value: str) -> str:
    return "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in value).strip("._-") or "job"


def make_job_dir(base_output_dir: Path, workbook_path: Path) -> Path:
    label = sanitize_label(f"{workbook_path.stem}_ns3_subtype_resistance_profile")
    job_dir = base_output_dir / label
    if job_dir.exists():
        shutil.rmtree(job_dir)
    job_dir.mkdir(parents=True, exist_ok=True)
    return job_dir


def format_freq(value: float) -> str:
    if value <= 0:
        return "0"
    if value >= 10.0:
        return f"{value:.0f}"
    if value >= 1.0:
        return f"{value:.1f}"
    return format(value, ".1g")


def variants_to_rich_text(variants: VariantCell) -> CellRichText | str:
    if not variants:
        return ""
    parts: list[str | TextBlock] = []
    for aa, pct in variants:
        parts.append(aa)
        parts.append(TextBlock(InlineFont(vertAlign="superscript"), pct))
    return CellRichText(*parts)


def format_coverage_range(values: list[int]) -> str:
    if not values:
        return "0-0"
    return f"{min(values)}-{max(values)}"


def load_consensus_by_gt(json_path: Path) -> dict[str, str]:
    rows = json.loads(json_path.read_text(encoding="utf-8"))
    consensus: dict[str, str] = {}
    for row in rows:
        name = str(row.get("name", ""))
        match = re.fullmatch(r"HCV([1-8])NS3", name)
        if match:
            consensus[match.group(1)] = str(row.get("refSequence", "")).strip().upper()
    return consensus


def load_subtype_profile_rows(
    workbook_path: Path,
) -> tuple[
    dict[str, dict[str, dict[int, list[tuple[str, float]]]]],
    dict[str, dict[str, int]],
    dict[str, dict[str, dict[int, int]]],
]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    profile_rows: dict[str, dict[str, dict[int, list[tuple[str, float]]]]] = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    subtype_counts: dict[str, dict[str, int]] = defaultdict(dict)
    position_coverage: dict[str, dict[str, dict[int, int]]] = defaultdict(lambda: defaultdict(dict))
    for sheet_name in wb.sheetnames:
        gt = sheet_name.replace("GT", "")
        ws = wb[sheet_name]
        next(ws.iter_rows(values_only=True))
        for row in ws.iter_rows(min_row=2, values_only=True):
            subtype = str(row[0])
            pos = int(row[1])
            denom = int(row[2])
            aa = str(row[3])
            pct = float(row[6])
            if pos in RESISTANCE_POSITIONS:
                profile_rows[gt][subtype][pos].append((aa, pct))
                position_coverage[gt][subtype][pos] = denom
            current = subtype_counts[gt].get(subtype, 0)
            if denom > current:
                subtype_counts[gt][subtype] = denom
    wb.close()
    return profile_rows, subtype_counts, position_coverage


def build_grid(
    profile_rows: dict[str, dict[str, dict[int, list[tuple[str, float]]]]],
    subtype_counts: dict[str, dict[str, int]],
    position_coverage: dict[str, dict[str, dict[int, int]]],
) -> list[list[GridCell]]:
    grid: list[list[GridCell]] = [[""] + [f"P{pos}" for pos in RESISTANCE_POSITIONS]]
    ordered_subtypes: list[tuple[str, str]] = []
    for gt in sorted(subtype_counts, key=int):
        for subtype in sorted(subtype_counts[gt]):
            ordered_subtypes.append((gt, subtype))

    for gt, subtype in ordered_subtypes:
        pos_variants: dict[int, VariantCell] = {}
        for pos in RESISTANCE_POSITIONS:
            variants = [
                (aa, format_freq(pct))
                for aa, pct in sorted(profile_rows[gt][subtype].get(pos, []), key=lambda item: (-item[1], item[0]))
                if aa not in EXCLUDED_AAS and pct >= FREQUENCY_THRESHOLD_PERCENT
            ]
            pos_variants[pos] = variants

        coverage_values = [position_coverage[gt][subtype].get(pos, 0) for pos in RESISTANCE_POSITIONS]
        row: list[GridCell] = [
            f"GT{gt}_{subtype} ({subtype_counts[gt][subtype]}, {format_coverage_range(coverage_values)})",
        ]
        for pos in RESISTANCE_POSITIONS:
            row.append(pos_variants[pos])
        grid.append(row)
    return grid


def write_excel(path: Path, grid: list[list[GridCell]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Subtype_Resistance_Profile"
    block_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    bold = Font(bold=True)

    for row_idx, row in enumerate(grid, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = variants_to_rich_text(value) if isinstance(value, list) else value
        first = row[0]
        if isinstance(first, str) and first.startswith("GT"):
            for cell in ws[row_idx]:
                cell.fill = block_fill
                cell.font = bold
        for cell in ws[row_idx]:
            cell.alignment = Alignment(horizontal="center")

    from openpyxl.utils import get_column_letter
    for col in range(1, len(RESISTANCE_POSITIONS) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12 if col > 1 else 18
    wb.save(path)


def main() -> int:
    args = parse_args()
    subtype_profile_workbook = Path(args.subtype_profile_workbook).expanduser()
    gt_aa_json = Path(args.gt_aa_json).expanduser()
    output_dir = Path(args.output_dir)
    script_temp_dir()

    profile_rows, subtype_counts, position_coverage = load_subtype_profile_rows(subtype_profile_workbook)
    grid = build_grid(profile_rows, subtype_counts, position_coverage)

    output_dir.mkdir(parents=True, exist_ok=True)
    excel_path = output_dir / "NS3_Subtype_RAS_Profiles.xlsx"
    write_excel(excel_path, grid)

    summary = {
        "excel": str(excel_path.resolve()),
        "positions": RESISTANCE_POSITIONS,
        "frequency_threshold_percent": FREQUENCY_THRESHOLD_PERCENT,
    }
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
